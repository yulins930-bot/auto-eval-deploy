# -*- coding: utf-8 -*-
"""
自动化评测 — 本地真实评测
上传 → LLM 理解/澄清 → 备份并规范化数据 → 单元测试 → 全量跑批
启动：python app.py        浏览器 http://127.0.0.1:5050
"""
from __future__ import annotations

import csv
import json
import os
import re
import shutil
import threading
import time
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import quote

from flask import Flask, Response, jsonify, request, send_from_directory

from vision_utils import (
    assess_job_vision,
    build_excel_image_index,
    detect_embedded_image_columns,
    resolve_student_answer_images,
    vision_prompt_suffix,
)
from platform_core import (
    append_vision_limit_note,
    apply_prompt_engineering,
    audit_clarified_output,
    describe_active_strategies,
    preview_prompt_engineering_append,
    backup_and_prepare,
    compare_row_ground_truths,
    finalize_field_mapping,
    compare_prediction,
    compute_column_stats,
    count_ground_truth_rows,
    ensure_prompt_essentials,
    extract_confidence,
    extract_prediction,
    format_row_ground_truth_summary,
    format_row_prediction_summary,
    format_per_column_match_line,
    infer_scoring_profile,
    list_ground_truth_columns,
    row_judgment_correct,
    load_clarify_followup_template,
    load_clarify_interpret_template,
    load_clarify_template,
    majority_prediction,
    render_clarify_followup_prompt,
    render_clarify_interpret_prompt,
    metrics_for_responses,
    normalize_strategies,
    parse_model_json,
    preview_from_table,
    render_clarify_prompt,
    render_row_prompt,
    resolve_row_ground_truth,
    row_to_dict,
    select_sample_rows,
    suggest_annotation_columns,
    suggest_core_input_columns,
    suggest_primary_content_columns,
    validate_field_mapping,
    validate_prompt_renders,
)

# ── 路径 ──────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DEMO_DIR = Path(__file__).resolve().parent
PROTOTYPE_DIR = ROOT / "prototype"
UPLOAD_DIR = DEMO_DIR / "uploads"
BACKUP_DIR = DEMO_DIR / "backups"
PREPARED_DIR = DEMO_DIR / "prepared"
TEMPLATES_DIR = DEMO_DIR / "templates"
TASK_REGISTRY_PATH = DEMO_DIR / "task_registry.json"
SNAPSHOT_DIR = DEMO_DIR / "snapshots"
for _d in (UPLOAD_DIR, BACKUP_DIR, PREPARED_DIR, SNAPSHOT_DIR):
    _d.mkdir(parents=True, exist_ok=True)

# ── Flask ─────────────────────────────────────────────
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

# ── 内存状态（重启即清空，DEMO 足够）──────────────────
JOBS: dict[str, dict] = {}   # job_id  → 上传元信息
RUNS: dict[str, dict] = {}   # run_id  → 跑批状态与结果
_GEMINI_LOCAL = threading.local()


def _rehydrate_job_from_disk(job_id: str) -> dict | None:
    """多 worker / 进程重启后，从 uploads 目录恢复 Job（内存 JOBS 不共享）。"""
    matches = sorted(
        UPLOAD_DIR.glob(f"{job_id}_*"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        return None
    path = matches[0]
    try:
        columns, data_rows = inspect_file(path)
        file_data = read_rows(path)
        col_stats = compute_column_stats(file_data["columns"], file_data["rows"])
    except Exception:
        return None
    ext = path.suffix.lower()
    file_kind = "xlsx" if ext == ".xlsx" else "csv"
    display_name = path.name[len(job_id) + 1 :] if path.name.startswith(f"{job_id}_") else path.name
    job: dict[str, Any] = {
        "job_id": job_id,
        "filename": display_name,
        "saved_name": path.name,
        "columns": columns,
        "data_rows": data_rows,
        "column_stats": col_stats,
        "path": str(path),
        "file_kind": file_kind,
        "embedded_image_columns": [],
    }
    if file_kind == "xlsx":
        try:
            job["embedded_image_columns"] = detect_embedded_image_columns(path, columns)
        except Exception:
            pass
    prepared = PREPARED_DIR / f"{path.stem}_prepared.csv"
    if prepared.is_file():
        job["prepared_path"] = str(prepared)
        job["status"] = "ready"
    return job


def _get_job(job_id: str | None) -> dict | None:
    if not job_id:
        return None
    job = JOBS.get(job_id)
    if job:
        return job
    job = _rehydrate_job_from_disk(job_id)
    if job:
        JOBS[job_id] = job
    return job


def _now_iso() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def _init_run(run_id: str, job_id: str, job: dict, cfg: dict) -> dict:
    """创建前端可追踪的运行记录。"""
    return {
        "run_id": run_id,
        "job_id": job_id,
        "filename": job.get("filename", ""),
        "task_description": cfg.get("task_description") or job.get("task_description") or "",
        "config": cfg,
        "status": "queued",
        "created_at": _now_iso(),
        "finished_at": None,
        "started_at": None,
        "cancel_requested": False,
        "progress": 0,
        "processed_in_model": 0,
        "total_rows": 0,
        "total_models": 0,
        "current_model": "",
        "current_model_idx": 0,
        "model_progress": [],
        "model_results": [],
        "evaluated_rows": None,
        "error": None,
    }


def _mark_run_finished(run: dict, status: str, error: str | None = None) -> None:
    run["status"] = status
    run["finished_at"] = _now_iso()
    if error:
        run["error"] = error


def _run_should_cancel(run: dict) -> bool:
    return bool(run.get("cancel_requested"))


def _response_at_index(responses: list, idx: int) -> dict:
    if idx < len(responses):
        r = responses[idx]
        if isinstance(r, dict):
            return r
    for r in responses:
        if isinstance(r, dict) and int(r.get("row") or 0) == idx + 1:
            return r
    return {}


def _estimate_run_seconds(job: dict, cfg: dict, model_specs: list[dict]) -> int | None:
    """启动前粗估总耗时（秒），供界面展示。"""
    try:
        file_data = read_rows(_resolve_data_path(job))
        rows = select_sample_rows(
            file_data.get("rows") or [],
            sample_percent=cfg.get("sample_percent"),
            max_rows=cfg.get("max_rows"),
        )
        n_rows = len(rows)
        n_models = max(1, len(model_specs))
        conc = max(1, min(30, int(cfg.get("concurrency") or 10)))
        repeat = 2 if cfg.get("stability_check") else 1
        strategies = normalize_strategies(cfg.get("strategies") or {})
        if strategies.get("repeat_runs"):
            repeat = max(repeat, int(strategies.get("repeat_count") or 2))
        vision = bool(cfg.get("vision_required") or job.get("vision_required"))
        base_sec = 18 if vision else 5
        units = n_rows * n_models * repeat
        if cfg.get("model_parallel", True) and n_models > 1:
            parallel = conc * min(n_models, int(cfg.get("model_parallelism") or 4))
        else:
            parallel = conc
        return max(20, int(units * base_sec / max(parallel, 1)))
    except Exception:
        return None


def _format_duration_sec(sec: float | None) -> str:
    if sec is None:
        return "—"
    s = max(0, int(round(sec)))
    if s < 60:
        return f"{s} 秒"
    m, r = divmod(s, 60)
    if m < 60:
        return f"{m} 分 {r} 秒"
    h, m = divmod(m, 60)
    return f"{h} 小时 {m} 分"


def _run_timing_fields(run: dict) -> dict[str, Any]:
    """根据已用时间与进度估算剩余/总耗时（秒）。"""
    started = run.get("started_at")
    progress = float(run.get("progress") or 0)
    elapsed = round(time.time() - started, 1) if started else 0.0
    hint = run.get("estimated_total_sec_hint")
    if progress > 0.001 and elapsed > 0:
        estimated_total = round(elapsed / progress, 1)
        estimated_remaining = max(0.0, round(estimated_total - elapsed, 1))
    elif hint:
        estimated_total = float(hint)
        estimated_remaining = max(0.0, round(estimated_total - elapsed, 1))
    else:
        estimated_total = None
        estimated_remaining = None
    return {
        "elapsed_sec": elapsed,
        "estimated_total_sec": estimated_total,
        "estimated_remaining_sec": estimated_remaining,
        "estimated_total_label": _format_duration_sec(estimated_total),
        "estimated_remaining_label": _format_duration_sec(estimated_remaining),
        "elapsed_label": _format_duration_sec(elapsed),
    }


def _finalize_model_result(
    mr: dict,
    *,
    total: int,
    gt_col: str | None,
    gt_cols: list[str],
    field_mapping: dict,
    cancelled: bool = False,
) -> dict:
    responses = mr.get("responses") or []
    if isinstance(responses, list) and len(responses) < total:
        responses = responses + [None] * (total - len(responses))
    mr["responses"] = responses
    mr["end_time"] = time.time()
    if mr.get("start_time"):
        mr["duration_sec"] = round(mr["end_time"] - mr["start_time"], 1)
    done_responses = [r for r in responses if isinstance(r, dict)]
    mr["errors"] = sum(1 for r in done_responses if r.get("status") == "error")
    if done_responses:
        mr["metrics"] = metrics_for_responses(
            done_responses,
            ground_truth_column=gt_col,
            ground_truth_columns=gt_cols,
            field_mapping=field_mapping,
        )
    else:
        mr["metrics"] = {}
    if cancelled:
        mr["cancelled"] = True
    return mr


def _run_history_entry(run: dict) -> dict:
    models = run.get("model_results") or []
    best_pct = None
    for mr in models:
        met = mr.get("metrics") or {}
        pct = met.get("accuracy_pct")
        if pct is not None and (best_pct is None or pct > best_pct):
            best_pct = pct
    return {
        "run_id": run.get("run_id"),
        "job_id": run.get("job_id"),
        "filename": run.get("filename") or (JOBS.get(run.get("job_id"), {}) or {}).get("filename", ""),
        "task_type": run.get("task_type") or (JOBS.get(run.get("job_id"), {}) or {}).get("task_type", ""),
        "task_description": run.get("task_description") or "",
        "status": run.get("status"),
        "created_at": run.get("created_at"),
        "finished_at": run.get("finished_at"),
        "progress": round(float(run.get("progress") or 0), 4),
        "total_rows": run.get("total_rows") or 0,
        "total_models": run.get("total_models") or 0,
        "best_accuracy_pct": best_pct,
        "snapshot_saved": bool(run.get("snapshot_path")),
        "error": run.get("error"),
    }


def _short_text(value: Any, limit: int = 500) -> str:
    text = "" if value is None else str(value).strip()
    return text if len(text) <= limit else text[:limit] + "…"


def _has_row_prediction(row: dict) -> bool:
    if row.get("predicted") is not None:
        return True
    by_col = row.get("ground_truth_by_column") or {}
    return any((d or {}).get("predicted") is not None for d in by_col.values())


def _build_harness_spec(job: dict, cfg: dict | None = None, fm: dict | None = None) -> dict:
    """把当前任务配置整理成可展示、可存档的自动化验证契约。"""
    cfg = cfg or {}
    clarified = job.get("clarified") or {}
    fm = finalize_field_mapping(fm or cfg.get("field_mapping") or job.get("field_mapping") or clarified.get("field_mapping") or {})
    gt_cols = list_ground_truth_columns(fm)
    core_inputs = fm.get("core_inputs") or ([fm.get("primary_content")] if fm.get("primary_content") else [])
    prompt_template = cfg.get("prompt_template") or job.get("prompt_template") or clarified.get("suggested_prompt") or ""
    output_schema = clarified.get("output_schema") or ""
    has_gt = bool(gt_cols)
    return {
        "task_spec": {
            "task_name": job.get("filename") or "未命名任务",
            "task_type": job.get("task_type") or clarified.get("task_type") or "未分类",
            "user_intent": job.get("task_description") or cfg.get("task_description") or "",
            "row_meaning": "每一行作为一次独立 AI 判断任务处理",
            "core_inputs": [c for c in core_inputs if c],
            "reference_columns": [c for c in [fm.get("reference_column"), fm.get("stem_column"), fm.get("rubric_column")] if c],
            "rubric_source": job.get("rubric_filename") or ("已粘贴/内置规则" if job.get("rubric_text") else ""),
        },
        "prompt_spec": {
            "output_format": job.get("output_format") or clarified.get("output_format") or "",
            "output_description": job.get("output_description") or clarified.get("output_description") or job.get("user_output_expectation") or "",
            "output_schema": output_schema,
            "prompt_chars": len(prompt_template),
            "strategies": describe_active_strategies(cfg.get("strategies") or job.get("strategies") or {}),
        },
        "eval_spec": {
            "has_ground_truth": has_gt,
            "ground_truth_columns": gt_cols,
            "primary_ground_truth": fm.get("primary_ground_truth") or fm.get("ground_truth"),
            "evaluation_mode": fm.get("evaluation_mode") or "auto",
            "comparison_unit": fm.get("comparison_unit"),
            "comparison_unit_label": fm.get("comparison_unit_label"),
            "row_headline": fm.get("row_headline"),
            "row_headline_label": fm.get("row_headline_label"),
            "judgment_metric_note": fm.get("judgment_metric_note"),
            "prediction_key_mapping": fm.get("prediction_key_mapping") or {},
            "no_label_method": "多模型一致率 / 分歧样本 / 人工抽查建议" if not has_gt else "",
            "readiness_thresholds": {
                "suggest_automate_accuracy_pct": 85,
                "suggest_semi_auto_accuracy_pct": 70,
                "min_parse_success_pct": 80,
                "min_api_success_pct": 80,
                "automate_uses_row_strict_when_both": True,
            },
        },
    }


def _build_failure_buckets(model_results: list[dict], gt_cols: list[str]) -> dict:
    buckets = {
        "api_failure": {"label": "API 失败", "count": 0, "examples": []},
        "parse_failure": {"label": "输出不可解析", "count": 0, "examples": []},
        "not_comparable": {"label": "不可比对", "count": 0, "examples": []},
        "ground_truth_conflict": {"label": "与人工标注冲突", "count": 0, "examples": []},
    }

    def add(kind: str, model_name: str, row: dict, reason: str) -> None:
        b = buckets[kind]
        b["count"] += 1
        if len(b["examples"]) < 5:
            b["examples"].append({
                "model": model_name,
                "row": row.get("row"),
                "reason": reason,
                "response": _short_text(row.get("response"), 160),
            })

    for mr in model_results:
        name = mr.get("model_name") or mr.get("name") or "模型"
        for row in mr.get("responses") or []:
            if not isinstance(row, dict):
                continue
            if row.get("status") != "ok":
                add("api_failure", name, row, row.get("response") or "接口调用失败")
                continue
            has_gt = row.get("ground_truth") is not None or bool(row.get("ground_truth_by_column"))
            has_pred = _has_row_prediction(row)
            if not has_pred:
                add("parse_failure", name, row, "模型有返回，但未抽取到可比对的预测字段")
            elif has_gt and row.get("correct") is None and gt_cols:
                add("not_comparable", name, row, "预测与标注存在，但当前规则无法判断对错")
            elif row.get("correct") is False:
                add("ground_truth_conflict", name, row, row.get("match_explanation") or "预测与人工标注不一致")
    return buckets


def _build_consensus_report(model_results: list[dict]) -> dict:
    """无标注时用多模型一致性提供自动化信心，不包装成准确率。"""
    real = [mr for mr in model_results if not mr.get("synthetic") and mr.get("responses")]
    total = max((len(mr.get("responses") or []) for mr in real), default=0)
    comparable = 0
    full_agree = 0
    majority_agree = 0
    disagreements = []
    for i in range(total):
        values = []
        row_no = i + 1
        for mr in real:
            rows = mr.get("responses") or []
            if i >= len(rows):
                continue
            row = rows[i]
            if row.get("status") != "ok" or row.get("predicted") is None:
                continue
            values.append(str(row.get("predicted")).strip())
            row_no = row.get("row") or row_no
        if len(values) < 2:
            continue
        comparable += 1
        counts: dict[str, int] = {}
        for v in values:
            counts[v] = counts.get(v, 0) + 1
        top_value, top_count = max(counts.items(), key=lambda kv: kv[1])
        if len(counts) == 1:
            full_agree += 1
        if top_count >= 2:
            majority_agree += 1
        if len(counts) > 1 and len(disagreements) < 10:
            disagreements.append({
                "row": row_no,
                "values": counts,
                "majority": top_value if top_count >= 2 else None,
            })
    return {
        "model_count": len(real),
        "rows_compared": comparable,
        "full_consensus_rows": full_agree,
        "majority_consensus_rows": majority_agree,
        "full_consensus_pct": round(full_agree / comparable * 100, 2) if comparable else None,
        "majority_consensus_pct": round(majority_agree / comparable * 100, 2) if comparable else None,
        "disagreement_examples": disagreements,
        "message": (
            "无人工标注时，该指标只表示模型之间是否同意，不代表真实准确率。"
            if comparable else "无人工标注且有效多模型预测不足，暂无法形成一致性判断。"
        ),
    }


def _build_sample_gate(run: dict, model_results: list[dict], readiness_report: dict, consensus: dict) -> dict:
    responses = [r for mr in model_results if not mr.get("synthetic") for r in (mr.get("responses") or [])]
    total = len(responses)
    api_ok = sum(1 for r in responses if r.get("status") == "ok")
    parseable = sum(1 for r in responses if r.get("status") == "ok" and _has_row_prediction(r))
    api_ok_pct = round(api_ok / total * 100, 2) if total else None
    parse_pct = round(parseable / api_ok * 100, 2) if api_ok else None
    blockers = []
    warnings = []
    if total and (api_ok_pct or 0) < 80:
        blockers.append("API 跑通率低于 80%，不建议全量。")
    if api_ok and (parse_pct or 0) < 80:
        blockers.append("模型输出解析率低于 80%，请先收紧 Prompt / 输出契约。")
    if readiness_report.get("has_ground_truth") and readiness_report.get("best_metric_pct") is not None:
        if readiness_report["best_metric_pct"] < 70:
            warnings.append("小样本判断正确率低于 70%，建议先优化 Prompt 或换模型。")
    elif not readiness_report.get("has_ground_truth"):
        if consensus.get("model_count", 0) < 2:
            warnings.append("无标注场景建议至少选择 2 个模型，用一致率辅助判断。")
        elif consensus.get("majority_consensus_pct") is not None and consensus["majority_consensus_pct"] < 70:
            warnings.append("多模型一致率偏低，建议先人工抽查分歧样本。")
    ready = not blockers
    return {
        "mode": (run.get("config") or {}).get("mode"),
        "ready_for_batch": ready,
        "api_ok_pct": api_ok_pct,
        "parse_success_pct": parse_pct,
        "blockers": blockers,
        "warnings": warnings,
        "recommendation": "可以进入全量评测" if ready else "请先修复阻断项再全量",
    }


def _build_readiness_report(
    *,
    has_ground_truth: bool,
    best_model: str | None,
    best_metric_pct: float | None,
    automate_metric_pct: float | None = None,
    summary: dict,
    consensus: dict,
    failure_buckets: dict,
) -> dict:
    risks = []
    next_actions = []
    evidence = []
    api_ok_pct = summary.get("api_ok_pct")
    parse_risk = sum(
        int((failure_buckets.get(k) or {}).get("count") or 0)
        for k in ("parse_failure", "not_comparable")
    )
    if api_ok_pct is not None:
        evidence.append(f"API 跑通率 {api_ok_pct}%")
    if best_model:
        evidence.append(f"当前最佳模型：{best_model}")

    auto_pct = automate_metric_pct if automate_metric_pct is not None else best_metric_pct
    row_strict_pct = summary.get("row_strict_accuracy_pct")
    seq_label = summary.get("sequence_units_label")
    cmp_unit = summary.get("comparison_unit")

    if has_ground_truth:
        if best_metric_pct is not None:
            if cmp_unit == "sequence_char" and seq_label:
                evidence.append(f"主指标·逐字一致 {best_metric_pct}%（{seq_label}）")
                if row_strict_pct is not None and row_strict_pct != best_metric_pct:
                    evidence.append(f"辅指标·整串全对 {row_strict_pct}%（自动化门槛）")
            else:
                evidence.append(f"最佳业务一致率 {best_metric_pct}%")
        if auto_pct is not None and auto_pct != best_metric_pct:
            evidence.append(f"自动化参考一致率 {auto_pct}%")
        if auto_pct is not None and auto_pct >= 85 and parse_risk == 0:
            verdict = "建议自动化"
            level = "high"
            next_actions.append("用更多样本复跑一次，并准备固化到专用脚本或正式工作流。")
        elif auto_pct is not None and auto_pct >= 70:
            verdict = "建议半自动"
            level = "medium"
            risks.append("准确率尚未达到直接放手自动化的水平，需要人工复核低置信或边界样本。")
            next_actions.append("优先查看错误样本 Top N，补充 few-shot 或细化评分标准后复跑。")
        elif best_metric_pct is None:
            verdict = "暂不建议自动化"
            level = "low"
            risks.append("本次没有得到可计算的业务准确率，可能是标注缺失或输出不可解析。")
            next_actions.append("回第 2 步确认标注列和 JSON 输出契约，再跑 5 行试验。")
        else:
            verdict = "暂不建议自动化"
            level = "low"
            if cmp_unit == "sequence_char" and best_metric_pct and best_metric_pct >= 50:
                risks.append(
                    "逐字有一定正确率，但整串全对率仍偏低；自动化请主要看整串全对/辅指标，勿只看逐字。"
                )
            else:
                risks.append("当前业务一致率偏低，直接自动化会带来较高误判风险。")
            next_actions.append("先调整 Prompt / 输出契约或更换模型，再用小样本复测。")
    else:
        maj = consensus.get("majority_consensus_pct")
        full = consensus.get("full_consensus_pct")
        if maj is not None:
            evidence.append(f"多模型多数一致率 {maj}%")
        if full is not None:
            evidence.append(f"多模型完全一致率 {full}%")
        if maj is not None and maj >= 85:
            verdict = "建议半自动"
            level = "medium"
            risks.append("当前没有人工标注，一致率不能等同于真实准确率。")
            next_actions.append("抽查多模型分歧样本和少量一致样本，沉淀第一版人工标注集。")
        else:
            verdict = "暂不建议自动化"
            level = "low"
            risks.append("缺少人工标注，且多模型一致性不足或样本不足。")
            next_actions.append("至少选择两个模型试跑，并人工标注 20-50 条代表性样本。")

    if parse_risk:
        risks.append(f"有 {parse_risk} 条输出不可解析或不可比对，说明输出契约还不够稳定。")
    if summary.get("avg_cost_cny_per_row") is not None:
        evidence.append(f"估算单条成本 ¥{summary['avg_cost_cny_per_row']}")
    if not next_actions:
        next_actions.append("保存本次配置快照，作为后续复跑和交接基线。")
    return {
        "verdict": verdict,
        "level": level,
        "has_ground_truth": has_ground_truth,
        "best_model": best_model,
        "best_metric_pct": best_metric_pct,
        "evidence": evidence,
        "risks": risks,
        "next_actions": next_actions,
    }


def _save_run_snapshot(run: dict, payload: dict) -> str:
    cfg = dict(run.get("config") or {})
    cfg.pop("api_key", None)
    cfg.pop("api_keys", None)
    snapshot = {
        "run_id": run.get("run_id"),
        "job_id": run.get("job_id"),
        "created_at": run.get("created_at"),
        "finished_at": run.get("finished_at"),
        "config": cfg,
        "harness_spec": payload.get("harness_spec"),
        "readiness_report": payload.get("readiness_report"),
        "sample_gate": payload.get("sample_gate"),
        "consensus_report": payload.get("consensus_report"),
        "summary": payload.get("summary"),
        "models": payload.get("models"),
    }
    path = SNAPSHOT_DIR / f"{run.get('run_id')}.json"
    path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    run["snapshot_path"] = str(path)
    return path.name

# ── 厂商（用户选 API 时用）────────────────────────────
VENDORS = {
    "google": {
        "name": "Google Gemini",
        "key_hint": "AIza…（Google AI Studio）",
        "env_var": "GEMINI_API_KEY",
    },
    "openai": {
        "name": "OpenAI",
        "key_hint": "sk-…（platform.openai.com）",
        "env_var": "OPENAI_API_KEY",
    },
    "doubao": {
        "name": "字节豆包 / 火山方舟",
        "key_hint": "方舟 API Key（控制台创建）",
        "env_var": "ARK_API_KEY",
    },
    "glm": {
        "name": "智谱 GLM",
        "key_hint": "glm 开头或控制台 API Key",
        "env_var": "ZHIPU_API_KEY",
    },
    "qwen": {
        "name": "阿里通义",
        "key_hint": "DashScope API Key",
        "env_var": "DASHSCOPE_API_KEY",
    },
    "deepseek": {
        "name": "DeepSeek",
        "key_hint": "sk-...（platform.deepseek.com）",
        "env_var": "DEEPSEEK_API_KEY",
    },
    "anthropic": {
        "name": "Claude / Anthropic",
        "key_hint": "sk-ant-...（console.anthropic.com）",
        "env_var": "ANTHROPIC_API_KEY",
        "env_alt_vars": ["CLAUDE_API_KEY"],
    },
}


def _demo_access_code() -> str:
    """可选演示口令：设置后，只有输入该口令才允许使用服务端环境变量 Key。"""
    return (os.environ.get("DEMO_ACCESS_CODE") or "").strip()


def _is_demo_access_code(value: str | None) -> bool:
    code = _demo_access_code()
    return bool(code and (value or "").strip() == code)


def _env_api_key_for_vendor(vendor_id: str, presented_value: str | None = "") -> str:
    """解析服务端托管 Key。若设置 DEMO_ACCESS_CODE，则必须输入口令才启用。"""
    vendor = VENDORS.get(vendor_id, {})
    env_var = vendor.get("env_var", "")
    if not env_var:
        return ""
    if _demo_access_code() and not _is_demo_access_code(presented_value):
        return ""
    for name in [env_var] + list(vendor.get("env_alt_vars") or []):
        key = (os.environ.get(name) or "").strip()
        if key:
            return key
    return ""


def _resolve_presented_api_key(vendor_id: str, presented_value: str | None) -> str:
    """用户可填真实 Key；也可填 DEMO_ACCESS_CODE，由后端换成环境变量 Key。"""
    value = (presented_value or "").strip()
    if _is_demo_access_code(value):
        return _env_api_key_for_vendor(vendor_id, value)
    return value or _env_api_key_for_vendor(vendor_id, value)

# ── 模型注册表（vendor 与澄清/跑批所用 provider 对应）────
MODELS = {
    "gemini-3.5-flash": {
        "vendor": "google", "provider": "gemini", "model_id": "gemini-3.5-flash",
        "name": "Gemini 3.5 Flash", "supports_vision": True,
    },
    "gemini-3.1-pro": {
        "vendor": "google", "provider": "gemini", "model_id": "gemini-3.1-pro-preview",
        "name": "Gemini 3.1 Pro", "supports_vision": True,
    },
    "gemini-3.1-flash-lite": {
        "vendor": "google", "provider": "gemini", "model_id": "gemini-3.1-flash-lite",
        "name": "Gemini 3.1 Flash-Lite", "supports_vision": True,
    },
    "gpt-5.5": {
        "vendor": "openai", "provider": "openai", "model_id": "gpt-5.5",
        "name": "GPT-5.5", "supports_vision": True,
    },
    "gpt-5.2": {
        "vendor": "openai", "provider": "openai", "model_id": "gpt-5.2",
        "name": "GPT-5.2", "supports_vision": True,
    },
    "doubao-seed-pro": {
        "vendor": "doubao", "provider": "openai_compat", "model_id": "doubao-seed-2-0-pro-260215",
        "name": "豆包 Seed Pro", "base_url": "https://ark.cn-beijing.volces.com/api/v3",
        "supports_vision": True, "supports_thinking": True, "thinking_default": "disabled",
    },
    "doubao-seed-2-1-pro": {
        "vendor": "doubao", "provider": "openai_compat", "model_id": "doubao-seed-2-1-pro-260628",
        "name": "豆包 Seed 2.1 Pro", "base_url": "https://ark.cn-beijing.volces.com/api/v3",
        "supports_vision": True, "supports_thinking": True, "thinking_default": "disabled",
    },
    "doubao-seed-lite": {
        "vendor": "doubao", "provider": "openai_compat", "model_id": "doubao-seed-2-0-lite-260215",
        "name": "豆包 Seed Lite", "base_url": "https://ark.cn-beijing.volces.com/api/v3",
        "supports_vision": True, "supports_thinking": False, "thinking_default": "disabled",
    },
    "doubao-seed-mini": {
        "vendor": "doubao", "provider": "openai_compat", "model_id": "doubao-seed-2-0-mini-260215",
        "name": "豆包 Seed Mini", "base_url": "https://ark.cn-beijing.volces.com/api/v3",
        "supports_vision": True, "supports_thinking": True, "thinking_default": "disabled",
    },
    "doubao-seed-1-6-flash": {
        "vendor": "doubao", "provider": "openai_compat", "model_id": "doubao-seed-1-6-flash-250715",
        "name": "豆包 Seed 1.6 Flash", "base_url": "https://ark.cn-beijing.volces.com/api/v3",
        "supports_vision": False, "supports_thinking": False,
    },
    "doubao-seed-1-6-thinking": {
        "vendor": "doubao", "provider": "openai_compat", "model_id": "doubao-seed-1-6-thinking-250715",
        "name": "豆包 Seed 1.6 Thinking", "base_url": "https://ark.cn-beijing.volces.com/api/v3",
        "supports_vision": False, "supports_thinking": True, "thinking_default": "enabled",
    },
    "glm-4.6v": {
        "vendor": "glm", "provider": "openai_compat", "model_id": "glm-4.6v",
        "name": "GLM-4.6V", "base_url": "https://open.bigmodel.cn/api/paas/v4",
        "supports_vision": True,
    },
    "qwen-flash": {
        "vendor": "qwen", "provider": "openai_compat", "model_id": "qwen-flash",
        "name": "Qwen Flash", "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "supports_vision": False,
    },
    "qwen-plus": {
        "vendor": "qwen", "provider": "openai_compat", "model_id": "qwen-plus",
        "name": "Qwen Plus", "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "supports_vision": False,
    },
    "qwen-max": {
        "vendor": "qwen", "provider": "openai_compat", "model_id": "qwen-max",
        "name": "Qwen Max", "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "supports_vision": False,
    },
    "qwen3-max-thinking": {
        "vendor": "qwen", "provider": "openai_compat", "model_id": "qwen3-max-thinking",
        "name": "Qwen3 Max Thinking", "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "supports_vision": False,
    },
    "qwen-vl-plus": {
        "vendor": "qwen", "provider": "openai_compat", "model_id": "qwen-vl-plus",
        "name": "Qwen VL Plus", "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "supports_vision": True,
    },
    "deepseek-chat": {
        "vendor": "deepseek", "provider": "openai_compat", "model_id": "deepseek-chat",
        "name": "DeepSeek Chat (V3)", "base_url": "https://api.deepseek.com/v1",
        "supports_vision": False,
    },
    "deepseek-reasoner": {
        "vendor": "deepseek", "provider": "openai_compat", "model_id": "deepseek-reasoner",
        "name": "DeepSeek Reasoner (R1)", "base_url": "https://api.deepseek.com/v1",
        "supports_vision": False,
    },
    "claude-opus-4-8": {
        "vendor": "anthropic", "provider": "anthropic", "model_id": "claude-opus-4-8",
        "name": "Claude Opus 4.8", "supports_vision": True,
    },
    "claude-sonnet-4-6": {
        "vendor": "anthropic", "provider": "anthropic", "model_id": "claude-sonnet-4-6",
        "name": "Claude Sonnet 4.6", "supports_vision": True,
    },
    "claude-sonnet-4-5": {
        "vendor": "anthropic", "provider": "anthropic", "model_id": "claude-sonnet-4-5-20250929",
        "name": "Claude Sonnet 4.5", "supports_vision": True,
    },
    "claude-haiku-4-5": {
        "vendor": "anthropic", "provider": "anthropic", "model_id": "claude-haiku-4-5-20251001",
        "name": "Claude Haiku 4.5", "supports_vision": True,
    },
}

USD_TO_CNY = 7.20

MODEL_PRICE_USD_PER_1M_TOKENS = {
    # 粗略估算，仅用于界面展示；真实费用以各厂商账单为准。
    "gemini-3.5-flash": {"input": 0.50, "output": 3.00},
    "gemini-3.1-pro": {"input": 2.00, "output": 12.00},
    "gemini-3.1-flash-lite": {"input": 0.25, "output": 1.50},
    "gpt-5.5": {"input": 5.00, "output": 30.00},
    "gpt-5.2": {"input": 1.75, "output": 14.00},
    "doubao-seed-2-0-pro-260215": {"input": 0.47, "output": 2.37},
    "doubao-seed-2-1-pro-260628": {"input": 0.83, "output": 4.17},
    "doubao-seed-2-0-lite-260215": {"input": 0.09, "output": 0.53},
    "doubao-seed-2-0-mini-260215": {"input": 0.03, "output": 0.31},
    "doubao-seed-1-6-flash-250715": {"input": 0.022, "output": 0.219},
    "doubao-seed-1-6-thinking-250715": {"input": 0.14, "output": 1.10},
    "glm-4.6v": {"input": 0.30, "output": 0.90},
    "qwen-flash": {"input": 0.10, "output": 0.40},
    "qwen-plus": {"input": 0.40, "output": 1.20},
    "qwen-max": {"input": 1.20, "output": 6.00},
    "qwen3-max-thinking": {"input": 0.78, "output": 3.90},
    "qwen-vl-plus": {"input": 0.40, "output": 1.20},
    "deepseek-chat": {"input": 0.27, "output": 1.10},
    "deepseek-reasoner": {"input": 0.55, "output": 2.19},
    "claude-opus-4-8": {"input": 5.00, "output": 25.00},
    "claude-sonnet-4-6": {"input": 3.00, "output": 15.00},
    "claude-sonnet-4-5-20250929": {"input": 3.00, "output": 15.00},
    "claude-haiku-4-5-20251001": {"input": 1.00, "output": 5.00},
}

PROVIDER_PRICE_USD_PER_1M_TOKENS = {
    "gemini": {"input": 0.50, "output": 3.00},
    "openai": {"input": 5.00, "output": 30.00},
    "openai_compat": {"input": 0.30, "output": 1.20},
    "anthropic": {"input": 5.00, "output": 25.00},
}

# Gemini API 旧展示名 -> 实际 model_id（避免 404）
GEMINI_MODEL_ID_ALIASES = {
    "gemini-3.1-pro": "gemini-3.1-pro-preview",
}

# 仅支持默认 temperature（传 0 会 HTTP 400），请求体中省略该字段
OPENAI_MODELS_NO_CUSTOM_TEMPERATURE = frozenset({
    "gpt-5.5",
    "gpt-5.2",
})


def _resolve_gemini_model_id(model_id: str) -> str:
    mid = (model_id or "").strip()
    return GEMINI_MODEL_ID_ALIASES.get(mid, mid)


def _openai_should_send_temperature(model_id: str, temperature: float) -> bool:
    mid = (model_id or "").strip()
    if mid in OPENAI_MODELS_NO_CUSTOM_TEMPERATURE:
        return False
    return True


def _models_for_api() -> dict:
    by_vendor: dict[str, list] = {v: [] for v in VENDORS}
    for key, m in MODELS.items():
        v = m.get("vendor", "google")
        by_vendor.setdefault(v, []).append({
            "key": key,
            "name": m["name"],
            "model_id": m["model_id"],
            "supports_vision": bool(m.get("supports_vision")),
            "supports_thinking": bool(m.get("supports_thinking")),
        })
    return {"vendors": VENDORS, "by_vendor": by_vendor}


def _vendor_to_model_spec(vendor: str, model_id: str, name: str | None = None) -> dict:
    """将厂商 + API model_id 转为 call_llm 所需结构。"""
    vid = (vendor or "google").lower()
    base_urls = {
        "doubao": "https://ark.cn-beijing.volces.com/api/v3",
        "glm": "https://open.bigmodel.cn/api/paas/v4",
        "qwen": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "deepseek": "https://api.deepseek.com/v1",
    }
    if vid == "google":
        provider = "gemini"
    elif vid == "openai":
        provider = "openai"
    elif vid == "anthropic":
        provider = "anthropic"
    else:
        provider = "openai_compat"
    return {
        "vendor": vid,
        "provider": provider,
        "model_id": model_id,
        "name": name or model_id,
        "base_url": base_urls.get(vid),
        "supports_vision": True,
        "supports_thinking": vid == "doubao",
    }


def resolve_model_specs(model_keys: list[str], custom_models: list[dict] | None = None) -> list[dict]:
    """合并预设模型 key 与用户自定义 model_id。"""
    specs: list[dict] = []
    seen: set[str] = set()

    for mk in model_keys or []:
        if mk == "__custom__":
            continue
        if mk not in MODELS:
            continue
        if mk in seen:
            continue
        seen.add(mk)
        info = dict(MODELS[mk])
        info["key"] = mk
        specs.append(info)

    for cm in custom_models or []:
        vendor = (cm.get("vendor") or "google").strip()
        model_id = (cm.get("model_id") or cm.get("model_name") or "").strip()
        if not model_id:
            continue
        dedupe = f"{vendor}:{model_id}"
        if dedupe in seen:
            continue
        seen.add(dedupe)
        info = _vendor_to_model_spec(vendor, model_id, cm.get("name") or model_id)
        info["key"] = f"custom:{dedupe}"
        specs.append(info)

    return specs


def vendors_in_specs(model_specs: list[dict]) -> list[str]:
    seen: list[str] = []
    for m in model_specs:
        v = m.get("vendor", "google")
        if v not in seen:
            seen.append(v)
    return seen


def resolve_api_keys(body: dict, model_specs: list[dict]) -> dict[str, str]:
    """按厂商解析 API Key：api_keys 字典 > 单一 api_key（仅单厂商时）> 环境变量/演示口令。"""
    default = (body.get("api_key") or "").strip()
    raw_keys = body.get("api_keys") or {}
    if not isinstance(raw_keys, dict):
        raw_keys = {}

    out: dict[str, str] = {}
    vendors_needed = vendors_in_specs(model_specs)
    single_vendor = len(vendors_needed) == 1

    for vid in vendors_needed:
        k = (raw_keys.get(vid) or "").strip()
        if k:
            k = _resolve_presented_api_key(vid, k)
        if not k and single_vendor and default:
            k = _resolve_presented_api_key(vid, default)
        if not k:
            k = _env_api_key_for_vendor(vid, "")
        if k:
            out[vid] = k
    return out


def validate_api_keys_for_specs(model_specs: list[dict], api_keys: dict[str, str]) -> str | None:
    missing = [v for v in vendors_in_specs(model_specs) if not api_keys.get(v)]
    if not missing:
        return None
    names = [VENDORS.get(v, {}).get("name", v) for v in missing]
    return "缺少以下厂商的 API Key：" + "、".join(names) + "。请在第 3 步为各厂商分别填写。"


def _save_clarify_model_prefs(job: dict, body: dict, model_key: str) -> None:
    job["clarify_model_key"] = model_key
    if model_key == "__custom__":
        job["clarify_custom_vendor"] = (
            body.get("custom_vendor") or body.get("vendor") or job.get("clarify_custom_vendor") or "google"
        )
        job["clarify_custom_model_id"] = (
            body.get("custom_model_id") or body.get("model_id") or job.get("clarify_custom_model_id") or ""
        )
    else:
        job.pop("clarify_custom_model_id", None)


def _resolve_clarify_model(body: dict, job: dict | None = None) -> tuple[dict, dict]:
    """澄清阶段：预设 key 或自定义 vendor + model_id；可从 job 回退（补充澄清时 Step1 控件可能不可见）。"""
    job = job or {}
    model_key = (body.get("model_key") or job.get("clarify_model_key") or "").strip()
    if model_key == "__custom__" or (model_key not in MODELS and job.get("clarify_custom_model_id")):
        vendor = (
            body.get("custom_vendor") or body.get("vendor")
            or job.get("clarify_custom_vendor") or job.get("clarify_vendor") or "google"
        ).strip()
        model_id = (
            body.get("custom_model_id") or body.get("model_id")
            or job.get("clarify_custom_model_id") or ""
        ).strip()
        if not model_id:
            raise ValueError("请填写自定义模型 ID（如 gemini-3.5-flash），或返回第 1 步重新选择模型")
        mi = _vendor_to_model_spec(vendor, model_id, body.get("custom_model_name") or model_id)
        mi["key"] = "__custom__"
        return mi, VENDORS.get(vendor, {})
    if not model_key or model_key not in MODELS:
        fallback = job.get("clarify_model_key") or "gemini-3.5-flash"
        if fallback in MODELS:
            model_key = fallback
        else:
            raise ValueError("请选择用于「理解需求」的厂商和模型，或使用自定义模型 ID")
    mi = dict(MODELS[model_key])
    mi["key"] = model_key
    return mi, VENDORS.get(mi.get("vendor", "google"), {})

# ═══════════════════════════════════════════════════════
# 文件解析
# ═══════════════════════════════════════════════════════

def _safe_name(raw: str) -> str:
    name = raw.strip().replace(" ", "_")
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    return name or "upload"


def _parse_xlsx_meta(path: Path) -> tuple[list[str], int]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = []
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(1, c).value
            headers.append(str(v).strip() if v is not None and str(v).strip() else f"列{c}")
        return headers, max(0, (ws.max_row or 1) - 1)
    finally:
        wb.close()


def _parse_csv_meta(path: Path) -> tuple[list[str], int]:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    if not lines:
        return [], 0
    reader = csv.reader([lines[0]])
    headers = next(reader)
    headers = [h.strip() if h else f"列{i+1}" for i, h in enumerate(headers)]
    return headers, max(0, len(lines) - 1)


def inspect_file(path: Path) -> tuple[list[str], int]:
    if path.suffix.lower() == ".csv":
        return _parse_csv_meta(path)
    return _parse_xlsx_meta(path)


def read_rows(path: Path) -> dict:
    """返回 {"columns": [...], "rows": [[v,v,...], ...]}"""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return _read_csv_rows(path)
    return _read_xlsx_rows(path)


def _read_xlsx_rows(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    cols = []
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(1, c).value
        cols.append(str(v).strip() if v is not None and str(v).strip() else f"列{c}")
    rows = []
    for r in range(2, (ws.max_row or 1) + 1):
        row = []
        for c in range(1, len(cols) + 1):
            v = ws.cell(r, c).value
            row.append("" if v is None else str(v))
        rows.append(row)
    wb.close()
    return {"columns": cols, "rows": rows}


def _read_csv_rows(path: Path) -> dict:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    if not lines:
        return {"columns": [], "rows": []}
    reader = csv.reader(lines)
    headers = next(reader, [])
    rows = list(reader)
    return {"columns": headers, "rows": rows}


def _compact_preview_rows(
    rows: list[dict[str, str]],
    *,
    max_rows: int = 3,
    max_cell_chars: int = 180,
) -> list[dict[str, str]]:
    """压缩澄清样本，减少 token 体积，提升首轮理解速度。"""
    out: list[dict[str, str]] = []
    for row in (rows or [])[:max_rows]:
        rr: dict[str, str] = {}
        for k, v in (row or {}).items():
            s = "" if v is None else str(v)
            if len(s) > max_cell_chars:
                s = s[:max_cell_chars] + "…"
            rr[str(k)] = s
        out.append(rr)
    return out


# ═══════════════════════════════════════════════════════
# LLM 调用
# ═══════════════════════════════════════════════════════

def call_llm(
    provider: str,
    model_id: str,
    api_key: str,
    prompt: str,
    temperature: float = 0,
    base_url: str | None = None,
    images: list[dict] | None = None,
    extra_body: dict | None = None,
) -> str:
    text, _usage = call_llm_with_usage(
        provider,
        model_id,
        api_key,
        prompt,
        temperature,
        base_url=base_url,
        images=images,
        extra_body=extra_body,
    )
    return text


def call_llm_with_usage(
    provider: str,
    model_id: str,
    api_key: str,
    prompt: str,
    temperature: float = 0,
    base_url: str | None = None,
    images: list[dict] | None = None,
    extra_body: dict | None = None,
) -> tuple[str, dict]:
    imgs = images or []
    if provider == "gemini":
        return _call_gemini_with_usage(model_id, api_key, prompt, temperature, images=imgs)
    if provider == "anthropic":
        return _call_anthropic_with_usage(model_id, api_key, prompt, images=imgs)
    if provider in ("openai", "openai_compat"):
        url = (base_url or "https://api.openai.com/v1") + "/chat/completions"
        return _call_openai_compat_with_usage(url, model_id, api_key, prompt, temperature, images=imgs, extra_body=extra_body)
    raise ValueError(f"不支持的 provider: {provider}")


def _usage_obj_to_dict(obj: Any) -> dict:
    if obj is None:
        return {}
    if isinstance(obj, dict):
        return obj
    out = {}
    for name in (
        "prompt_tokens",
        "completion_tokens",
        "total_tokens",
        "cached_tokens",
        "reasoning_tokens",
        "thought_tokens",
    ):
        v = getattr(obj, name, None)
        if v is not None:
            out[name] = v
    for nested in ("prompt_tokens_details", "completion_tokens_details", "output_tokens_details"):
        v = getattr(obj, nested, None)
        if v is not None:
            out[nested] = _usage_obj_to_dict(v)
    return out


def model_supports_vision(model_info: dict) -> bool:
    key = model_info.get("key") or ""
    if key.startswith("custom:"):
        return model_info.get("supports_vision", True)
    if key in MODELS:
        return bool(MODELS[key].get("supports_vision"))
    return bool(model_info.get("supports_vision", True))


def _call_gemini(
    model_id: str,
    api_key: str,
    prompt: str,
    temperature: float,
    *,
    images: list[dict] | None = None,
) -> str:
    text, _usage = _call_gemini_with_usage(model_id, api_key, prompt, temperature, images=images)
    return text


def _get_gemini_client(api_key: str):
    from google import genai

    client = getattr(_GEMINI_LOCAL, "client", None)
    cached_key = getattr(_GEMINI_LOCAL, "api_key", None)
    if client is None or cached_key != api_key:
        _GEMINI_LOCAL.client = genai.Client(api_key=api_key)
        _GEMINI_LOCAL.api_key = api_key
    return _GEMINI_LOCAL.client


def _call_gemini_with_usage(
    model_id: str,
    api_key: str,
    prompt: str,
    temperature: float,
    *,
    images: list[dict] | None = None,
) -> tuple[str, dict]:
    import time as time_mod

    from google.genai import types

    model_id = _resolve_gemini_model_id(model_id)
    client = _get_gemini_client(api_key)
    contents: list[Any] = []
    for img in images or []:
        contents.append(
            types.Part.from_bytes(data=img["bytes"], mime_type=img.get("mime") or "image/jpeg")
        )
    contents.append(prompt)
    last_err: Exception | None = None
    for attempt in range(3):
        try:
            resp = client.models.generate_content(
                model=model_id,
                contents=contents if len(contents) > 1 else prompt,
                config=types.GenerateContentConfig(temperature=temperature),
            )
            usage = _usage_obj_to_dict(getattr(resp, "usage_metadata", None))
            if usage:
                usage.setdefault("provider_usage_source", "gemini_usage_metadata")
            return resp.text or "", usage
        except Exception as e:
            last_err = e
            msg = str(e).lower()
            retryable = any(x in msg for x in ("429", "quota", "rate", "timeout", "temporarily", "unavailable", "503", "502", "504"))
            if attempt < 2 and retryable:
                time_mod.sleep(1.5 * (attempt + 1))
                continue
            raise
    if last_err:
        raise last_err
    raise RuntimeError("Gemini 接口调用失败")


def _call_openai_compat(
    url: str,
    model_id: str,
    api_key: str,
    prompt: str,
    temperature: float,
    *,
    images: list[dict] | None = None,
    extra_body: dict | None = None,
) -> str:
    text, _usage = _call_openai_compat_with_usage(
        url,
        model_id,
        api_key,
        prompt,
        temperature,
        images=images,
        extra_body=extra_body,
    )
    return text


def _call_openai_compat_with_usage(
    url: str,
    model_id: str,
    api_key: str,
    prompt: str,
    temperature: float,
    *,
    images: list[dict] | None = None,
    extra_body: dict | None = None,
) -> tuple[str, dict]:
    import requests as req
    import time

    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    if images:
        import base64 as b64mod

        parts: list[dict] = [{"type": "text", "text": prompt}]
        for img in images:
            mime = img.get("mime") or "image/jpeg"
            b64 = b64mod.standard_b64encode(img["bytes"]).decode("ascii")
            parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}"},
            })
        user_content: Any = parts
    else:
        user_content = prompt
    body: dict[str, Any] = {
        "model": model_id,
        "messages": [{"role": "user", "content": user_content}],
    }
    if _openai_should_send_temperature(model_id, temperature):
        body["temperature"] = temperature
    if extra_body:
        body.update(extra_body)
    last_err: Exception | None = None
    for attempt in range(3):
        try:
            r = req.post(url, json=body, headers=headers, timeout=(60, 240))
            if r.status_code == 400 and "temperature" in (r.text or "").lower() and "temperature" in body:
                body_retry = {k: v for k, v in body.items() if k != "temperature"}
                r = req.post(url, json=body_retry, headers=headers, timeout=(60, 240))
            r.raise_for_status()
            payload = r.json()
            usage = payload.get("usage") or {}
            if isinstance(usage, dict):
                usage.setdefault("provider_usage_source", "chat_completions_usage")
            return payload["choices"][0]["message"]["content"], usage if isinstance(usage, dict) else {}
        except req.HTTPError as e:
            last_err = _http_error_detail(e)
            code = e.response.status_code if e.response is not None else 0
            if code in (429, 502, 503, 504) and attempt < 2:
                time.sleep(2.0 * (attempt + 1))
                continue
            raise last_err from e
        except (req.Timeout, req.ConnectionError, req.exceptions.ChunkedEncodingError) as e:
            last_err = e
            time.sleep(1.5 * (attempt + 1))
        except req.RequestException as e:
            last_err = e
            msg = str(e).lower()
            if attempt < 2 and ("connection" in msg or "retries" in msg or "timeout" in msg):
                time.sleep(1.5 * (attempt + 1))
                continue
            raise
    if last_err:
        raise last_err
    raise RuntimeError("OpenAI 兼容接口调用失败")


def _call_anthropic_with_usage(
    model_id: str,
    api_key: str,
    prompt: str,
    *,
    images: list[dict] | None = None,
) -> tuple[str, dict]:
    import base64 as b64mod
    import requests as req
    import time

    content: list[dict[str, Any]] = []
    for img in images or []:
        mime = img.get("mime") or "image/jpeg"
        b64 = b64mod.standard_b64encode(img["bytes"]).decode("ascii")
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": mime,
                "data": b64,
            },
        })
    content.append({"type": "text", "text": prompt})

    body = {
        "model": model_id,
        "max_tokens": 4096,
        "messages": [{"role": "user", "content": content}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    last_err: Exception | None = None
    for attempt in range(3):
        try:
            r = req.post("https://api.anthropic.com/v1/messages", json=body, headers=headers, timeout=(60, 240))
            r.raise_for_status()
            payload = r.json()
            text_parts = []
            for item in payload.get("content") or []:
                if isinstance(item, dict) and item.get("type") == "text":
                    text_parts.append(item.get("text") or "")
            usage = payload.get("usage") or {}
            if isinstance(usage, dict):
                usage.setdefault("prompt_tokens", usage.get("input_tokens", 0))
                usage.setdefault("completion_tokens", usage.get("output_tokens", 0))
                usage.setdefault("total_tokens", (usage.get("input_tokens") or 0) + (usage.get("output_tokens") or 0))
                usage.setdefault("provider_usage_source", "anthropic_messages_usage")
            return "\n".join(text_parts).strip(), usage if isinstance(usage, dict) else {}
        except req.HTTPError as e:
            last_err = _http_error_detail(e)
            code = e.response.status_code if e.response is not None else 0
            if code in (429, 502, 503, 504) and attempt < 2:
                time.sleep(2.0 * (attempt + 1))
                continue
            raise last_err from e
        except (req.Timeout, req.ConnectionError, req.exceptions.ChunkedEncodingError) as e:
            last_err = e
            time.sleep(1.5 * (attempt + 1))
        except req.RequestException as e:
            last_err = e
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise
    if last_err:
        raise last_err
    raise RuntimeError("Anthropic 接口调用失败")


# ═══════════════════════════════════════════════════════
# Prompt 构造
# ═══════════════════════════════════════════════════════

def build_row_prompt(task_desc: str, columns: list[str], row_values: list[str]) -> str:
    """兼容旧逻辑：无确认 prompt 时回退为任务说明 + 行数据。"""
    pairs = []
    for col, val in zip(columns, row_values):
        v = val.strip() if val else ""
        if v and v.lower() != "nan":
            pairs.append(f"  {col}: {v}")
    row_text = "\n".join(pairs) if pairs else "  （本行数据为空）"
    return f"""## 任务说明
{task_desc}

## 当前行数据
{row_text}

请严格根据「任务说明」处理上述数据，直接输出结果。"""


def _resolve_data_path(job: dict) -> Path:
    if job.get("prepared_path") and Path(job["prepared_path"]).is_file():
        return Path(job["prepared_path"])
    return Path(job["path"])


def _discover_files_on_disk() -> list[dict]:
    """扫描工作区内常见 sample 路径（供任务库展示）。"""
    patterns = [
        "**/sample.xlsx",
        "**/sample_clean.xlsx",
        "**/prior_test_100.csv",
        "**/prior_test_500_crop.csv",
        "**/test_cases.csv",
        "原任务/**/data/*.csv",
    ]
    found: list[dict] = []
    seen: set[str] = set()
    for pat in patterns:
        for p in ROOT.glob(pat):
            if not p.is_file():
                continue
            rel = str(p.relative_to(ROOT)).replace("\\", "/")
            if rel in seen or "副本" in rel:
                continue
            seen.add(rel)
            found.append({"path": rel, "size_kb": round(p.stat().st_size / 1024, 1)})
    return sorted(found, key=lambda x: x["path"])[:80]


def _provider_model_defaults(provider: str, model_name: str) -> tuple[str, str, str | None]:
    """返回 call_llm 用的 provider, model_id, base_url。"""
    p = (provider or "gemini").lower()
    if p in ("google", "gemini"):
        return "gemini", model_name or "gemini-3.5-flash", None
    if p == "openai":
        return "openai", model_name or "gpt-5.5", None
    if p in ("doubao", "volc", "ark"):
        return "openai_compat", model_name or "doubao-seed-2-0-lite-260215", "https://ark.cn-beijing.volces.com/api/v3"
    if p == "glm":
        return "openai_compat", model_name or "glm-4.6v", "https://open.bigmodel.cn/api/paas/v4"
    if p == "qwen":
        return "openai_compat", model_name or "qwen-flash", "https://dashscope.aliyuncs.com/compatible-mode/v1"
    if p == "deepseek":
        return "openai_compat", model_name or "deepseek-chat", "https://api.deepseek.com/v1"
    return "gemini", model_name or "gemini-3.5-flash", None


def _rough_tokens(text: str) -> int:
    # 中英文混合场景下用字符数粗估 token，用于成本趋势展示。
    return max(1, int(len(text or "") / 3))


def _model_price_usd(model_info: dict) -> dict:
    model_id = (model_info.get("model_id") or "").strip()
    provider = model_info.get("provider") or "gemini"
    return (
        MODEL_PRICE_USD_PER_1M_TOKENS.get(model_id)
        or PROVIDER_PRICE_USD_PER_1M_TOKENS.get(provider)
        or PROVIDER_PRICE_USD_PER_1M_TOKENS["openai_compat"]
    )


def estimate_call_cost_usd(model_info: dict, prompt: str, response: str) -> float:
    price = _model_price_usd(model_info)
    input_tokens = _rough_tokens(prompt)
    output_tokens = _rough_tokens(response)
    return round((input_tokens * price["input"] + output_tokens * price["output"]) / 1_000_000, 6)


def estimate_call_cost_cny(model_info: dict, prompt: str, response: str) -> float:
    return round(estimate_call_cost_usd(model_info, prompt, response) * USD_TO_CNY, 6)


def _int_usage(usage: dict, *paths: str) -> int:
    for path in paths:
        cur: Any = usage or {}
        ok = True
        for part in path.split("."):
            if isinstance(cur, dict) and part in cur:
                cur = cur[part]
            else:
                ok = False
                break
        if ok:
            try:
                return int(cur or 0)
            except (TypeError, ValueError):
                continue
    return 0


def calculate_call_cost(model_info: dict, prompt: str, response: str, usage: dict | None = None) -> dict:
    """
    优先用厂商返回的 usage 计费；豆包 thinking 产生的 token 会体现在 completion/usage 中。
    无 usage 时才回退到字符粗估，并明确标记 estimated=True。
    """
    price = _model_price_usd(model_info)
    usage = usage or {}
    prompt_tokens = _int_usage(usage, "prompt_tokens", "input_tokens")
    completion_tokens = _int_usage(usage, "completion_tokens", "output_tokens")
    reasoning_tokens = _int_usage(
        usage,
        "reasoning_tokens",
        "thought_tokens",
        "completion_tokens_details.reasoning_tokens",
        "output_tokens_details.reasoning_tokens",
    )
    used_provider_usage = prompt_tokens > 0 or completion_tokens > 0
    if not used_provider_usage:
        prompt_tokens = _rough_tokens(prompt)
        completion_tokens = _rough_tokens(response)
    usd = (prompt_tokens * price["input"] + completion_tokens * price["output"]) / 1_000_000
    return {
        "usd": round(usd, 8),
        "cny": round(usd * USD_TO_CNY, 8),
        "estimated": not used_provider_usage,
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "reasoning_tokens": reasoning_tokens,
        "usage": usage,
    }


def build_cost_estimation_info(model_specs: list[dict] | None = None) -> dict:
    specs = model_specs or []
    prices = []
    seen = set()
    for m in specs:
        model_id = m.get("model_id") or ""
        if not model_id or model_id in seen:
            continue
        seen.add(model_id)
        p = _model_price_usd(m)
        prices.append({
            "model": m.get("name") or model_id,
            "model_id": model_id,
            "input_cny_per_1m_tokens": round(p["input"] * USD_TO_CNY, 4),
            "output_cny_per_1m_tokens": round(p["output"] * USD_TO_CNY, 4),
        })
    return {
        "currency": "CNY",
        "exchange_rate_usd_to_cny": USD_TO_CNY,
        "token_estimation": "优先使用 API 返回的 usage；接口不返回 usage 时才按 len(text)/3 粗估",
        "formula": "单次成本 = (usage.prompt_tokens * 输入单价 + usage.completion_tokens * 输出单价) / 1,000,000；thinking token 按厂商 usage 计入输出侧",
        "note": "模型单价按公开 API 口径折算人民币；实际费用以厂商账单为准。标记为 estimated 的行表示该接口未返回 usage。",
        "prices": prices,
    }


def normalize_result_analysis(parsed: dict, raw_text: str = "") -> dict:
    """保证结果解读可读，并让策略字段稳定对应第二步控件。"""
    if not isinstance(parsed, dict) or parsed.get("parse_error"):
        return {
            "verdict": "解读完成（原始文本）",
            "verdict_reason": "",
            "next_action": "请根据下方摘要人工判断",
            "summary": (raw_text or "")[:1500],
            "findings": [],
            "suggestions": [],
            "strategy_recommendations": normalize_strategies({"enabled": False}),
            "prompt_revision": "",
        }
    out = {
        "verdict": str(parsed.get("verdict") or "已生成评测解读").strip(),
        "verdict_reason": str(parsed.get("verdict_reason") or "").strip(),
        "next_action": str(parsed.get("next_action") or "").strip(),
        "summary": str(parsed.get("summary") or "").strip(),
        "findings": [],
        "suggestions": [],
        "strategy_recommendations": normalize_strategies(parsed.get("strategy_recommendations") or {}),
        "prompt_revision": str(parsed.get("prompt_revision") or "").strip(),
    }
    findings = parsed.get("findings") or []
    if isinstance(findings, list):
        out["findings"] = [str(x).strip() for x in findings if str(x).strip()][:5]
    suggestions = parsed.get("suggestions") or []
    if isinstance(suggestions, list):
        cleaned = []
        for item in suggestions[:5]:
            if isinstance(item, dict):
                cleaned.append({
                    "title": str(item.get("title") or "建议").strip(),
                    "detail": str(item.get("detail") or "").strip(),
                    "priority": str(item.get("priority") or "medium").strip(),
                })
            elif str(item).strip():
                cleaned.append({"title": "建议", "detail": str(item).strip(), "priority": "medium"})
        out["suggestions"] = cleaned
    if not out["summary"] and out["findings"]:
        out["summary"] = "；".join(out["findings"][:2])
    if not out["next_action"] and out["suggestions"]:
        out["next_action"] = out["suggestions"][0]["detail"] or out["suggestions"][0]["title"]
    return out


def _prediction_signature(parsed: dict, raw_text: str) -> str:
    """用于稳定性：忽略解释/置信度等非判定字段，只比较核心输出。"""
    if isinstance(parsed, dict) and not parsed.get("parse_error"):
        drop = {"confidence", "reason", "explanation", "理由", "解释", "raw"}
        core = {str(k): v for k, v in parsed.items() if str(k) not in drop and not str(k).startswith("_")}
        try:
            return json.dumps(core, ensure_ascii=False, sort_keys=True)
        except Exception:
            pass
    pred = extract_prediction(parsed, raw_text)
    return "" if pred is None else str(pred).strip()


def _build_ensemble_vote_result(
    model_results: list[dict],
    field_mapping: dict,
    gt_cols: list[str],
) -> dict | None:
    """多模型逐行投票：生成一个合成 model_result，不额外调用 API。"""
    ok_models = [mr for mr in model_results if mr.get("responses")]
    if len(ok_models) < 2:
        return None
    total = max((len(mr.get("responses") or []) for mr in ok_models), default=0)
    if total <= 0:
        return None
    profile = infer_scoring_profile([], field_mapping)
    responses: list[dict] = []
    for i in range(total):
        row_votes = []
        source_rows = []
        for mr in ok_models:
            rows = mr.get("responses") or []
            if i >= len(rows):
                continue
            r = rows[i]
            if r.get("status") != "ok":
                continue
            source_rows.append(r)
            if r.get("predicted") is not None:
                row_votes.append(r.get("predicted"))
        if not source_rows:
            responses.append({
                "row": i + 1,
                "response": "no successful model response for voting",
                "status": "error",
                "time_sec": 0,
            })
            continue

        by_col: dict[str, dict] = {}
        for col in gt_cols:
            values = []
            gt_val = None
            for r in source_rows:
                detail = (r.get("ground_truth_by_column") or {}).get(col) or {}
                if gt_val is None and detail.get("ground_truth") is not None:
                    gt_val = detail.get("ground_truth")
                if detail.get("predicted") is not None:
                    values.append(detail.get("predicted"))
            voted = majority_prediction(values) if len(values) >= 2 else None
            by_col[col] = {
                "ground_truth": gt_val,
                "predicted": voted,
                "correct": compare_prediction(voted, gt_val),
            }

        if not any((d or {}).get("predicted") is not None for d in by_col.values()):
            voted_row = majority_prediction(row_votes) if len(row_votes) >= 2 else None
            primary = field_mapping.get("primary_ground_truth") or field_mapping.get("ground_truth")
            gt_val = (source_rows[0] or {}).get("ground_truth")
            if primary and primary in gt_cols:
                by_col[primary] = {
                    "ground_truth": gt_val,
                    "predicted": voted_row,
                    "correct": compare_prediction(voted_row, gt_val),
                }

        if not any((d or {}).get("predicted") is not None for d in by_col.values()):
            responses.append({
                "row": i + 1,
                "response": "有效预测模型不足，未形成多模型投票",
                "status": "error",
                "time_sec": 0,
            })
            continue

        correct = row_judgment_correct(by_col, field_mapping)
        pred_display = format_row_prediction_summary(by_col, profile)
        if not pred_display:
            pred_display = majority_prediction(row_votes)
        gt_display = (source_rows[0] or {}).get("ground_truth")
        responses.append({
            "row": i + 1,
            "response": "多模型投票：" + json.dumps(
                {c: (by_col.get(c) or {}).get("predicted") for c in gt_cols},
                ensure_ascii=False,
            ),
            "parsed": {"vote": True},
            "predicted": pred_display,
            "ground_truth": gt_display,
            "ground_truth_by_column": by_col,
            "correct": correct,
            "match_explanation": format_per_column_match_line(by_col, gt_cols),
            "status": "ok",
            "time_sec": 0,
            "api_calls": 0,
            "estimated_cost_usd": 0,
            "estimated_cost_cny": 0,
        })
    return {
        "model_key": "__ensemble_vote__",
        "model_name": "多模型投票",
        "model_id": "majority-vote",
        "responses": responses,
        "errors": sum(1 for r in responses if r.get("status") == "error"),
        "duration_sec": 0,
        "synthetic": True,
        "metrics": metrics_for_responses(
            responses,
            ground_truth_column=field_mapping.get("primary_ground_truth") or field_mapping.get("ground_truth"),
            ground_truth_columns=gt_cols,
            field_mapping=field_mapping,
        ),
    }


# ═══════════════════════════════════════════════════════
# 后台跑批线程
# ═══════════════════════════════════════════════════════

def _call_llm_once(
    model_info: dict,
    api_key: str,
    prompt: str,
    temperature: float,
    *,
    images: list[dict] | None = None,
    extra_body: dict | None = None,
) -> tuple[str, float, dict]:
    t0 = time.time()
    resp_text, usage = call_llm_with_usage(
        model_info["provider"],
        model_info["model_id"],
        api_key,
        prompt,
        temperature,
        base_url=model_info.get("base_url"),
        images=images,
        extra_body=extra_body,
    )
    return resp_text, round(time.time() - t0, 2), usage


def _http_error_detail(err: Exception) -> Exception:
    """把 HTTP 4xx/5xx 响应体写入异常信息，便于区分 401 鉴权与 400 参数错误。"""
    import requests as req

    if not isinstance(err, req.HTTPError) or err.response is None:
        return err
    code = err.response.status_code
    msg = ""
    try:
        payload = err.response.json()
        if isinstance(payload, dict):
            err_obj = payload.get("error")
            if isinstance(err_obj, dict):
                msg = str(err_obj.get("message") or err_obj.get("code") or "")
            else:
                msg = str(payload.get("message") or payload.get("error") or "")
    except Exception:
        msg = (err.response.text or "")[:300]
    hint = ""
    if code == 401:
        hint = "（请检查该厂商 API Key 是否有效、是否已开通对应模型；Anthropic 需 sk-ant- 开头密钥）"
    elif code == 403:
        hint = "（账号可能未开通该模型或欠费）"
    detail = (msg or err.response.reason or "HTTP error").strip()
    return RuntimeError(f"HTTP {code}: {detail}{hint}")


def _doubao_thinking_extra(model_info: dict, cfg: dict) -> dict | None:
    if model_info.get("vendor") != "doubao" or not model_info.get("supports_thinking"):
        return None
    mode = (cfg.get("doubao_thinking") or model_info.get("thinking_default") or "disabled").strip().lower()
    if mode not in ("enabled", "disabled", "auto"):
        mode = "disabled"
    # 关闭时不传 thinking 字段，避免部分型号返回 400/401
    if mode == "disabled":
        return None
    thinking: dict[str, Any] = {"type": mode}
    if mode == "enabled":
        budget = cfg.get("doubao_thinking_budget_tokens")
        try:
            budget_i = int(budget or 0)
        except (TypeError, ValueError):
            budget_i = 0
        if budget_i > 0:
            thinking["budget_tokens"] = max(1024, min(32768, budget_i))
    return {"thinking": thinking}


def _process_one_row(
    row_index: int,
    row_values: list[str],
    *,
    columns: list[str],
    cfg: dict,
    model_info: dict,
    api_key: str,
) -> dict:
    task_desc = cfg.get("task_description", "")
    temperature = cfg.get("temperature", 0)
    prompt_template = cfg.get("prompt_template") or ""
    field_mapping = finalize_field_mapping(cfg.get("field_mapping") or {})
    strategies = normalize_strategies(cfg.get("strategies"))

    row_dict = row_to_dict(columns, row_values)
    prompt_extra: dict[str, Any] = {}
    rubric_text = (cfg.get("rubric_text") or "").strip()
    if rubric_text:
        prompt_extra["rubric_content"] = rubric_text
    if prompt_template:
        prompt = render_row_prompt(
            prompt_template, row_dict, field_mapping, extra_vars=prompt_extra or None
        )
    else:
        prompt = build_row_prompt(task_desc, columns, row_values)

    file_kind = cfg.get("file_kind") or "csv"
    image_path = cfg.get("image_source_path") or cfg.get("source_path") or ""
    embedded_cols = cfg.get("embedded_image_columns") or []
    answer_images = resolve_student_answer_images(
        row_index,
        row_dict,
        field_mapping,
        columns,
        image_source_path=image_path,
        file_kind=file_kind,
        embedded_image_columns=embedded_cols,
        excel_image_index=cfg.get("excel_image_index"),
    )
    use_vision = bool(answer_images)
    extra_body = _doubao_thinking_extra(model_info, cfg)
    prompt, _essential_notes = ensure_prompt_essentials(
        prompt,
        row_dict,
        field_mapping,
        vision_has_images=use_vision,
        extra_vars=prompt_extra or None,
    )
    prompt = apply_prompt_engineering(prompt, strategies, field_mapping)
    profile = infer_scoring_profile(columns, field_mapping)

    if use_vision:
        prompt = (prompt or "").rstrip() + vision_prompt_suffix()
    elif cfg.get("vision_required"):
        prompt = (prompt or "").rstrip() + (
            "\n\n【注意】本行未能加载主要输入图片（URL 不可达或 Excel 单元格无内嵌图），"
            "请结合其余文本字段判断。"
        )
    else:
        prompt = append_vision_limit_note(prompt, row_dict, field_mapping)

    stability_check = bool(cfg.get("stability_check"))
    repeat_n = strategies.get("repeat_count", 2) if strategies.get("repeat_runs") else 1
    if stability_check:
        repeat_n = max(repeat_n, 2)
    preds: list[Any] = []
    signatures: list[str] = []
    confidences: list[float] = []
    last_resp = ""
    total_sec = 0.0
    estimated_cost = 0.0
    estimated_cost_cny = 0.0
    cost_estimated = False
    prompt_tokens_total = 0
    completion_tokens_total = 0
    reasoning_tokens_total = 0
    usage_samples: list[dict] = []

    try:
        for _ in range(repeat_n):
            resp_text, elapsed, usage = _call_llm_once(
                model_info, api_key, prompt, temperature, images=answer_images or None, extra_body=extra_body
            )
            total_sec += elapsed
            cost_info = calculate_call_cost(model_info, prompt, resp_text, usage)
            estimated_cost += cost_info["usd"]
            estimated_cost_cny += cost_info["cny"]
            cost_estimated = cost_estimated or bool(cost_info["estimated"])
            prompt_tokens_total += int(cost_info.get("prompt_tokens") or 0)
            completion_tokens_total += int(cost_info.get("completion_tokens") or 0)
            reasoning_tokens_total += int(cost_info.get("reasoning_tokens") or 0)
            if usage:
                usage_samples.append(usage)
            last_resp = resp_text
            parsed = parse_model_json(resp_text)
            signatures.append(_prediction_signature(parsed, resp_text))
            pred = extract_prediction(parsed, resp_text)
            if pred is not None:
                preds.append(pred)
            conf = extract_confidence(parsed, resp_text)
            if conf is not None:
                confidences.append(conf)

        parsed = parse_model_json(last_resp)
        predicted = majority_prediction(preds) if len(preds) > 1 else (
            preds[0] if preds else extract_prediction(parsed, last_resp)
        )
        primary_gt = field_mapping.get("primary_ground_truth") or field_mapping.get("ground_truth")
        ground_truth = resolve_row_ground_truth(
            row_dict, field_mapping, columns, gt_column=primary_gt
        )
        by_col = compare_row_ground_truths(
            parsed, last_resp, row_dict, field_mapping, columns, default_predicted=predicted
        )
        confidence = sum(confidences) / len(confidences) if confidences else extract_confidence(parsed, last_resp)
        gt_cols = list_ground_truth_columns(field_mapping)

        low_conf = False
        if strategies.get("confidence_filter") and confidence is not None:
            if confidence < float(strategies.get("confidence_threshold") or 0.8):
                low_conf = True

        correct = None
        if not low_conf:
            correct = row_judgment_correct(by_col, field_mapping)
        pred_display = format_row_prediction_summary(by_col, profile) or predicted
        gt_display = format_row_ground_truth_summary(row_dict, field_mapping, columns) or ground_truth
        match_line = format_per_column_match_line(by_col, gt_cols)
        stability_consistent = None
        if stability_check and len(signatures) >= 2:
            non_empty = [s for s in signatures if s]
            stability_consistent = bool(non_empty) and len(set(non_empty)) == 1
        return {
            "row": row_index + 1,
            "response": last_resp,
            "parsed": parsed,
            "predicted": pred_display,
            "ground_truth": gt_display,
            "ground_truth_by_column": by_col,
            "correct": correct,
            "match_explanation": match_line,
            "vision_used": use_vision,
            "is_multi_point": profile.get("is_multi_point"),
            "confidence": confidence,
            "low_confidence": low_conf,
            "repeat_runs": repeat_n if repeat_n > 1 else None,
            "api_calls": repeat_n,
            "estimated_cost_usd": round(estimated_cost, 6),
            "estimated_cost_cny": round(estimated_cost_cny, 6),
            "cost_estimated": cost_estimated,
            "prompt_tokens": prompt_tokens_total,
            "completion_tokens": completion_tokens_total,
            "reasoning_tokens": reasoning_tokens_total,
            "usage_samples": usage_samples[:2],
            "stability_check": stability_check,
            "stability_consistent": stability_consistent,
            "status": "ok",
            "time_sec": round(total_sec, 2),
        }
    except Exception as e:
        return {
            "row": row_index + 1,
            "response": str(e),
            "status": "error",
            "api_calls": repeat_n if 'repeat_n' in locals() else 1,
            "estimated_cost_usd": round(estimated_cost, 6) if 'estimated_cost' in locals() else 0,
            "estimated_cost_cny": round(estimated_cost_cny, 6) if 'estimated_cost_cny' in locals() else 0,
            "cost_estimated": cost_estimated if 'cost_estimated' in locals() else True,
            "time_sec": total_sec,
        }


def _run_rows_for_models(
    run: dict,
    *,
    columns: list[str],
    rows: list[list[str]],
    cfg: dict,
    model_specs: list[dict],
) -> None:
    """对给定行列表跑多个模型；行级并发。"""
    total = len(rows)
    api_keys = cfg.get("api_keys") or {}
    default_api_key = cfg.get("api_key", "")
    field_mapping = finalize_field_mapping(cfg.get("field_mapping") or {})
    gt_col = field_mapping.get("primary_ground_truth") or field_mapping.get("ground_truth")
    gt_cols = list_ground_truth_columns(field_mapping)
    concurrency = max(1, min(30, int(cfg.get("concurrency") or 10)))

    run["total_rows"] = total
    run["total_models"] = len(model_specs)
    run["status"] = "running"
    run["started_at"] = time.time()
    run["cancel_requested"] = False
    run["concurrency"] = concurrency
    run["evaluated_rows"] = rows
    excel_image_index = cfg.get("excel_image_index")
    if excel_image_index is None and (cfg.get("file_kind") or "") == "xlsx":
        src = cfg.get("image_source_path") or cfg.get("source_path") or ""
        if src:
            try:
                excel_image_index = build_excel_image_index(Path(src), columns)
                cfg["excel_image_index"] = excel_image_index
            except Exception:
                excel_image_index = None
    run["model_progress"] = [
        {
            "idx": i,
            "name": m.get("name") or m.get("model_id") or f"模型{i + 1}",
            "model_id": m.get("model_id", ""),
            "done": 0,
            "total": total,
            "status": "queued",
            "errors": 0,
        }
        for i, m in enumerate(model_specs)
    ]
    if cfg.get("vision_required"):
        bad = [m["name"] for m in model_specs if not model_supports_vision(m)]
        if bad:
            _mark_run_finished(run, "error", (
                "本任务核心输入为图片，须使用支持多模态（视觉）的模型。"
                f"以下模型未标记为支持看图：{', '.join(bad)}"
            ))
            return
    progress_lock = threading.Lock()
    global_done = [0]
    grand_total = max(len(model_specs) * total, 1)
    model_parallel = cfg.get("model_parallel", True)
    model_workers = max(1, min(len(model_specs), int(cfg.get("model_parallelism") or 4)))

    def _run_single_model(mi: int, model_info: dict) -> tuple[int, dict]:
        if _run_should_cancel(run):
            mr = {
                "model_key": model_info.get("key", ""),
                "model_name": model_info["name"],
                "model_id": model_info.get("model_id", ""),
                "provider": model_info.get("provider", ""),
                "vendor": model_info.get("vendor", ""),
                "responses": [None] * total,
                "errors": 0,
                "start_time": time.time(),
            }
            return mi, _finalize_model_result(
                mr, total=total, gt_col=gt_col, gt_cols=gt_cols,
                field_mapping=field_mapping, cancelled=True,
            )
        vendor = model_info.get("vendor", "google")
        api_key = api_keys.get(vendor) or default_api_key
        if not api_key:
            vname = VENDORS.get(vendor, {}).get("name", vendor)
            raise ValueError(f"缺少 {vname} 的 API Key")

        with progress_lock:
            run["current_model"] = model_info["name"]
            run["current_model_idx"] = mi
            if mi < len(run.get("model_progress") or []):
                run["model_progress"][mi]["status"] = "running"
        mr = {
            "model_key": model_info.get("key", ""),
            "model_name": model_info["name"],
            "model_id": model_info.get("model_id", ""),
            "provider": model_info.get("provider", ""),
            "vendor": model_info.get("vendor", ""),
            "responses": [None] * total,
            "errors": 0,
            "start_time": time.time(),
        }
        model_done = [0]

        def _run_index(i: int, row_values: list[str]) -> tuple[int, dict]:
            item = _process_one_row(
                i, row_values, columns=columns, cfg=cfg, model_info=model_info, api_key=api_key
            )
            with progress_lock:
                model_done[0] += 1
                global_done[0] += 1
                run["processed_in_model"] = model_done[0]
                run["progress"] = global_done[0] / grand_total
                if mi < len(run.get("model_progress") or []):
                    run["model_progress"][mi]["done"] = model_done[0]
                    if item.get("status") == "error":
                        run["model_progress"][mi]["errors"] += 1
            return i, item

        cancelled = False
        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            futures = [pool.submit(_run_index, i, rv) for i, rv in enumerate(rows)]
            try:
                for fut in as_completed(futures):
                    if _run_should_cancel(run):
                        cancelled = True
                        for pending in futures:
                            pending.cancel()
                        break
                    i, item = fut.result()
                    mr["responses"][i] = item
                    if item.get("status") == "error":
                        mr["errors"] += 1
            finally:
                if cancelled:
                    pool.shutdown(wait=False, cancel_futures=True)

        mr = _finalize_model_result(
            mr,
            total=total,
            gt_col=gt_col,
            gt_cols=gt_cols,
            field_mapping=field_mapping,
            cancelled=cancelled,
        )
        with progress_lock:
            if mi < len(run.get("model_progress") or []):
                run["model_progress"][mi]["status"] = "cancelled" if cancelled else "done"
                run["model_progress"][mi]["done"] = model_done[0]
        if cancelled:
            return mi, mr
        return mi, mr

    results_by_idx: list[dict | None] = [None] * len(model_specs)
    if model_parallel and len(model_specs) > 1:
        with ThreadPoolExecutor(max_workers=model_workers) as model_pool:
            futs = [
                model_pool.submit(_run_single_model, mi, model_info)
                for mi, model_info in enumerate(model_specs)
            ]
            for fut in as_completed(futs):
                if _run_should_cancel(run):
                    for pending in futs:
                        pending.cancel()
                    break
                try:
                    mi, mr = fut.result()
                    results_by_idx[mi] = mr
                except Exception:
                    pass
                if _run_should_cancel(run):
                    break
    else:
        for mi, model_info in enumerate(model_specs):
            if _run_should_cancel(run):
                break
            _, mr = _run_single_model(mi, model_info)
            results_by_idx[mi] = mr
            if _run_should_cancel(run):
                break

    all_model_results = [r for r in results_by_idx if r is not None]
    run["model_results"] = all_model_results
    if _run_should_cancel(run):
        _mark_run_finished(run, "cancelled", "用户已中止评测，以下为已完成部分的结果")
    else:
        _mark_run_finished(run, "done")


def _apply_strategies_to_job(job: dict, body: dict) -> dict:
    if "strategies" in body:
        job["strategies"] = normalize_strategies(body["strategies"])
    return normalize_strategies(job.get("strategies"))


def _run_config_from_body(job: dict, body: dict, model_specs: list[dict], *, mode: str) -> dict:
    strategies = _apply_strategies_to_job(job, body)
    sample_percent = body.get("sample_percent")
    max_rows_raw = body.get("max_rows")
    max_rows = None
    if max_rows_raw is not None and str(max_rows_raw).isdigit():
        max_rows = int(max_rows_raw)
    if sample_percent is not None:
        try:
            sample_percent = float(sample_percent)
        except (TypeError, ValueError):
            sample_percent = None
    fm = finalize_field_mapping(body.get("field_mapping") or job.get("field_mapping") or {})
    file_kind = job.get("file_kind") or (
        "xlsx" if str(job.get("path", "")).lower().endswith(".xlsx") else "csv"
    )
    embedded_cols = job.get("embedded_image_columns") or []
    sample_rows: list[list[str]] = []
    try:
        src = Path(job["path"])
        if src.is_file():
            sample_rows = (read_rows(src).get("rows") or [])[:20]
    except Exception:
        pass
    vision_info = assess_job_vision(
        fm,
        file_kind=file_kind,
        embedded_image_columns=embedded_cols,
        columns=job.get("columns") or [],
        sample_rows=sample_rows,
    )
    job["vision_required"] = vision_info["vision_required"]
    fm["vision_required"] = vision_info["vision_required"]
    fm["primary_content_mode"] = vision_info["primary_content_mode"]
    return {
        "task_description": (body.get("task_description") or job.get("task_description") or "").strip(),
        "api_key": (body.get("api_key") or "").strip(),
        "api_keys": resolve_api_keys(body, model_specs),
        "models": body.get("models") or [],
        "custom_models": body.get("custom_models") or [],
        "model_specs": model_specs,
        "max_rows": max_rows,
        "sample_percent": sample_percent,
        "concurrency": int(body.get("concurrency") or 10),
        "model_parallel": body.get("model_parallel", True) is not False,
        "model_parallelism": int(body.get("model_parallelism") or 4),
        "temperature": float(body.get("temperature", 0)),
        "prompt_template": body.get("prompt_template") or job.get("prompt_template", ""),
        "field_mapping": fm,
        "output_format": job.get("output_format", ""),
        "strategies": strategies,
        "stability_check": bool(body.get("stability_check")),
        "doubao_thinking": (body.get("doubao_thinking") or "disabled").strip(),
        "doubao_thinking_budget_tokens": int(body.get("doubao_thinking_budget_tokens") or 0),
        "mode": mode,
        "vision_required": bool(vision_info["vision_required"]),
        "file_kind": job.get("file_kind") or "csv",
        "image_source_path": job.get("path") or "",
        "embedded_image_columns": job.get("embedded_image_columns") or [],
        "rubric_text": (job.get("rubric_text") or "").strip(),
    }


def _apply_field_mapping_to_job(job: dict, body: dict) -> dict:
    """跑批前合并前端传来的最新列映射（第二步下拉框）。"""
    fm_in = body.get("field_mapping")
    if not fm_in:
        return job.get("field_mapping") or {}
    columns = job.get("columns") or []
    field_mapping, warnings = validate_field_mapping(columns, fm_in)
    job["field_mapping"] = field_mapping
    if warnings:
        job["mapping_warnings"] = warnings
    return field_mapping


def _batch_worker(run_id: str) -> None:
    run = RUNS[run_id]
    try:
        job = _get_job(run.get("job_id"))
        if not job:
            _mark_run_finished(run, "error", "任务已失效，请重新上传并准备数据")
            return
        cfg = run["config"]
        file_data = read_rows(_resolve_data_path(job))
        columns = file_data["columns"]
        all_rows = file_data["rows"]
        rows = select_sample_rows(
            all_rows,
            sample_percent=cfg.get("sample_percent"),
            max_rows=cfg.get("max_rows"),
        )
        specs = cfg.get("model_specs") or resolve_model_specs(
            cfg.get("models", ["gemini-3.5-flash"]),
            cfg.get("custom_models"),
        )
        _run_rows_for_models(run, columns=columns, rows=rows, cfg=cfg, model_specs=specs)
    except Exception:
        _mark_run_finished(run, "error", traceback.format_exc())


# ═══════════════════════════════════════════════════════
# Flask 路由
# ═══════════════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory(PROTOTYPE_DIR, "index.html")


@app.route("/api/health")
def health():
    has_key = bool(os.environ.get("GEMINI_API_KEY"))
    pillow_ok = False
    try:
        from openpyxl.drawing.image import PILImage

        pillow_ok = bool(PILImage)
    except Exception:
        pillow_ok = False
    return jsonify({
        "ok": True,
        "has_env_key": has_key,
        "pillow_ok": pillow_ok,
        "excel_embedded_supported": pillow_ok,
        "flow": "upload→clarify→prepare→sample→batch",
    })


@app.route("/api/models")
def list_models():
    """供前端：先选厂商 → 再选模型 → 再填对应 API Key。"""
    return jsonify(_models_for_api())


@app.route("/api/evaluation_modes")
def list_evaluation_modes():
    """供第二步选择「评测 / 汇总方式」，与数据集类型解耦。"""
    from evaluation import evaluation_modes_for_ui

    return jsonify({"modes": evaluation_modes_for_ui()})


@app.route("/api/evaluation_plugins")
def list_evaluation_plugins_api():
    """列出可注册的自定义行级评测插件（内置 + evaluation_plugins_user）。"""
    from evaluation_plugins import list_plugins

    return jsonify({"plugins": list_plugins()})


@app.route("/api/tasks")
def list_tasks():
    registry = {}
    if TASK_REGISTRY_PATH.is_file():
        registry = json.loads(TASK_REGISTRY_PATH.read_text(encoding="utf-8"))
    discovered = _discover_files_on_disk()
    tasks = []
    for t in registry.get("tasks", []):
        hints = t.get("sample_hints", [])
        available = [h for h in hints if (ROOT / h.replace("/", os.sep)).is_file()]
        tasks.append({**t, "samples_available": available})
    return jsonify({"tasks": tasks, "discovered_files": discovered})


@app.route("/api/runs")
def list_runs():
    """最近评测记录（内存态，服务重启后清空）。"""
    items = sorted(
        (_run_history_entry(r) for r in RUNS.values()),
        key=lambda x: x.get("created_at") or "",
        reverse=True,
    )
    return jsonify({"runs": items[:100]})


@app.route("/api/load_sample", methods=["POST"])
def load_sample():
    """从工作区已有 sample 路径复制到 uploads，便于试跑原任务。"""
    body = request.get_json(silent=True) or {}
    rel = (body.get("path") or "").strip().replace("\\", "/")
    if not rel or ".." in rel:
        return jsonify({"error": "非法路径"}), 400
    src = ROOT / rel
    if not src.is_file():
        return jsonify({"error": f"文件不存在: {rel}"}), 404
    job_id = uuid.uuid4().hex[:10]
    safe = f"{job_id}_{src.name}"
    dst = UPLOAD_DIR / safe
    shutil.copy2(src, dst)
    columns, data_rows = inspect_file(dst)
    file_data = read_rows(dst)
    col_stats = compute_column_stats(file_data["columns"], file_data["rows"])
    ext = dst.suffix.lower()
    file_kind = "xlsx" if ext == ".xlsx" else "csv"
    embedded_cols: list[str] = []
    if file_kind == "xlsx":
        try:
            embedded_cols = detect_embedded_image_columns(dst, columns)
        except Exception:
            embedded_cols = []
    vision_hint = None
    if embedded_cols:
        vision_hint = (
            f"检测到 Excel 内嵌图片列：{'、'.join(embedded_cols)}。"
            "跑批时将把图片送入多模态模型，请选用支持看图的模型。"
        )
    JOBS[job_id] = {
        "job_id": job_id,
        "filename": src.name,
        "saved_name": safe,
        "columns": columns,
        "data_rows": data_rows,
        "column_stats": col_stats,
        "path": str(dst),
        "source_hint": rel,
        "file_kind": file_kind,
        "embedded_image_columns": embedded_cols,
    }
    return jsonify({
        "job_id": job_id,
        "filename": src.name,
        "columns": columns,
        "data_rows": data_rows,
        "column_count": len(columns),
        "column_stats": col_stats,
        "from": rel,
        "file_kind": file_kind,
        "embedded_image_columns": embedded_cols,
        "vision_hint": vision_hint,
    })


@app.route("/api/clarify", methods=["POST"])
def clarify():
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "请先上传或加载样本"}), 400
    sensitive_status = (body.get("sensitive_status") or "").strip()
    if sensitive_status == "unmasked":
        return jsonify({"error": "当前数据标记为未脱敏，不能进入第 2 步 LLM 理解需求。请先脱敏或改用测试样本。"}), 400
    if sensitive_status:
        job["sensitive_status"] = sensitive_status
    desc = (body.get("task_description") or body.get("description") or "").strip()
    if not desc:
        return jsonify({"error": "请填写任务/数据集说明"}), 400
    output_expect = (body.get("user_output_expectation") or "").strip()
    if not output_expect:
        return jsonify({"error": "请说明「你希望 AI 每行输出什么」"}), 400

    try:
        mi, vendor_info = _resolve_clarify_model(body, job)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    vendor = mi.get("vendor", "google")
    model_key = mi.get("key", "")
    _save_clarify_model_prefs(job, body, model_key)
    rubric_text = (body.get("rubric_text") or job.get("rubric_text") or "").strip()
    if body.get("rubric_text") is not None:
        job["rubric_text"] = rubric_text[:50000]
    api_key = _resolve_presented_api_key(vendor, body.get("api_key"))
    if not api_key:
        return jsonify({
            "error": f"请填写 {vendor_info.get('name', vendor)} 的 API Key 或演示口令（格式类似 {vendor_info.get('key_hint', '')}）",
        }), 400

    provider, model_id, base_url = mi["provider"], mi["model_id"], mi.get("base_url")

    path = Path(job["path"])
    file_data = read_rows(path)
    columns, preview, total = preview_from_table(
        file_data["columns"], file_data["rows"], job["data_rows"], n=5
    )
    preview_for_llm = _compact_preview_rows(preview, max_rows=3, max_cell_chars=180)
    col_stats = compute_column_stats(columns, file_data["rows"])
    annotation_hints = suggest_annotation_columns(
        columns, file_data["rows"], columns_stats=col_stats
    )
    primary_hints = suggest_core_input_columns(columns, file_data["rows"])
    tmpl = load_clarify_template(TEMPLATES_DIR)
    prompt = render_clarify_prompt(
        tmpl,
        task_description=desc,
        user_output_expectation=output_expect,
        file_path=str(path),
        columns_list=columns,
        preview_rows=preview_for_llm,
        annotation_column_hints=annotation_hints,
        primary_content_hints=primary_hints,
        rubric_text=job.get("rubric_text") or "",
    )
    try:
        raw = call_llm(provider, model_id, api_key, prompt, 0.1, base_url)
    except Exception as e:
        return jsonify({"error": f"澄清调用失败（{vendor_info.get('name', vendor)} / {mi['name']}）: {e}"}), 500
    parsed = parse_model_json(raw)
    parsed, repaired = _repair_suggested_prompt_with_llm(
        parsed,
        provider=provider,
        model_id=model_id,
        api_key=api_key,
        base_url=base_url,
        task_description=desc,
        user_output_expectation=output_expect,
        columns=columns,
    )
    clarified, _ = _clarified_from_parsed(parsed, columns)
    if repaired:
        clarified["mapping_warnings"] = list(clarified.get("mapping_warnings") or []) + ["首轮未产出 Prompt，已由模型自动补写。"]
    clarified["annotation_column_candidates"] = annotation_hints
    job["clarified"] = clarified
    job["clarifying_questions_snapshot"] = clarified.get("clarifying_questions") or []
    job["task_description"] = desc
    job["user_output_expectation"] = output_expect
    job["preview_rows"] = preview
    job["columns"] = columns
    job["column_stats"] = col_stats
    job["clarify_vendor"] = vendor
    job["status"] = "clarifying"
    return jsonify({
        "status": "clarifying",
        "job_id": job_id,
        "clarified": clarified,
        "columns": columns,
        "column_stats": col_stats,
        "data_rows": total,
        "preview_rows": preview,
        "model_used": {
            "vendor": vendor,
            "vendor_name": vendor_info.get("name", vendor),
            "model_key": model_key,
            "model_name": mi["name"],
        },
    })


def _clarified_from_parsed(parsed: dict, columns: list[str]) -> tuple[dict, list[str]]:
    raw_mapping = parsed.get("field_mapping", {}) or {}
    field_mapping, mapping_warnings = validate_field_mapping(columns, raw_mapping)
    field_mapping, audit_warnings = audit_clarified_output(
        parsed, field_mapping, columns=columns
    )
    mapping_warnings = list(mapping_warnings) + audit_warnings
    if not str(parsed.get("suggested_prompt") or "").strip():
        parsed["suggested_prompt"] = _build_fallback_suggested_prompt(parsed, field_mapping)
        mapping_warnings.append("模型未返回可用 Prompt，平台已自动生成一版可编辑草案。")
    sas = parsed.get("student_answer_source")
    if isinstance(sas, dict) and not isinstance(sas.get("column"), str):
        sas = None
    return {
        "understood_task": parsed.get("understood_task", ""),
        "student_answer_source": sas if isinstance(sas, dict) else None,
        "task_type": parsed.get("task_type", ""),
        "output_format": parsed.get("output_format", ""),
        "output_description": parsed.get("output_description", ""),
        "output_schema": parsed.get("output_schema", ""),
        "field_mapping": field_mapping,
        "suggested_prompt": parsed.get("suggested_prompt", ""),
        "clarifying_questions": parsed.get("clarifying_questions", []) or [],
        "confidence": parsed.get("confidence", 0),
        "raw_parse_error": parsed.get("parse_error"),
        "mapping_warnings": mapping_warnings,
    }, mapping_warnings


def _repair_suggested_prompt_with_llm(
    parsed: dict,
    *,
    provider: str,
    model_id: str,
    api_key: str,
    base_url: str | None,
    task_description: str,
    user_output_expectation: str,
    columns: list[str],
    field_mapping_hint: dict | None = None,
) -> tuple[dict, bool]:
    """
    模型澄清结果缺少 suggested_prompt 时，先让同模型补写一版。
    返回 (parsed, 是否成功补写)。
    """
    if str(parsed.get("suggested_prompt") or "").strip():
        return parsed, False
    try:
        repair_prompt = (
            "你是评测平台的 Prompt 生成器。下面是一次需求澄清结果，但 suggested_prompt 为空。\n"
            "请只返回 JSON：{\"suggested_prompt\":\"...\"}。\n\n"
            f"任务描述：{task_description or '(空)'}\n"
            f"期望输出：{user_output_expectation or '(空)'}\n"
            f"真实列名：{json.dumps(columns, ensure_ascii=False)}\n"
            f"字段映射草稿：{json.dumps(field_mapping_hint or parsed.get('field_mapping') or {}, ensure_ascii=False)}\n"
            f"澄清结果JSON：{json.dumps(parsed, ensure_ascii=False)}\n\n"
            "要求：\n"
            "1) Prompt 首块必须有核心输入占位（如 {{ primary_input }} 或 {{ 列名 }}）\n"
            "2) 若存在人工标注列，输出要求写明 JSON 键名需与标注列一致\n"
            "3) 不要输出 markdown，不要解释，只输出 JSON 对象\n"
        )
        raw = call_llm(provider, model_id, api_key, repair_prompt, 0.1, base_url)
        repaired = parse_model_json(raw)
        sp = (repaired.get("suggested_prompt") if isinstance(repaired, dict) else "") or ""
        if str(sp).strip():
            out = dict(parsed or {})
            out["suggested_prompt"] = str(sp).strip()
            return out, True
    except Exception:
        pass
    return parsed, False


def _build_fallback_suggested_prompt(parsed: dict, field_mapping: dict) -> str:
    """当澄清模型未返回 suggested_prompt 时，按映射生成可编辑的默认 Prompt。"""
    fm = finalize_field_mapping(field_mapping or {})
    primary = (fm.get("primary_content") or "").strip()
    core_inputs = [str(c).strip() for c in (fm.get("core_inputs") or []) if str(c).strip()]
    if primary and primary not in core_inputs:
        core_inputs.insert(0, primary)
    gt_cols = list_ground_truth_columns(fm)
    output_desc = (parsed.get("output_description") or "").strip()
    output_schema = (parsed.get("output_schema") or "").strip()

    if not core_inputs:
        core_block = "## 核心输入\n{{ primary_input }}\n"
    elif len(core_inputs) == 1:
        c = core_inputs[0]
        core_block = f"## 核心输入（{c}）\n{{{{ {c} }}}}\n"
    else:
        lines = ["## 核心输入"]
        for c in core_inputs:
            lines.append(f"### {c}")
            lines.append(f"{{{{ {c} }}}}")
            lines.append("")
        core_block = "\n".join(lines).rstrip() + "\n"

    if gt_cols:
        keys = "、".join(gt_cols)
        output_rule = (
            "## 输出要求\n"
            "请只输出一行 JSON（不要 markdown、不要多余解释）。\n"
            f"JSON 键名必须与以下人工标注列完全一致：{keys}。\n"
            "若某列无法判断，请给出 null，不要省略该键。\n"
        )
    else:
        output_rule = (
            "## 输出要求\n"
            "请只输出一行 JSON（不要 markdown、不要多余解释）。\n"
            "字段名要稳定，便于后续批量解析与统计。\n"
        )

    desc_block = f"## 任务目标\n{output_desc}\n" if output_desc else ""
    schema_block = f"## 输出结构示例\n{output_schema}\n" if output_schema else ""

    return (
        f"{core_block}\n"
        "## 评分标准（若有）\n{{ rubric_content }}\n\n"
        f"{desc_block}"
        f"{output_rule}\n"
        f"{schema_block}"
    ).strip() + "\n"


@app.route("/api/clarify_followup", methods=["POST"])
def clarify_followup():
    """用户填写补充说明后，再次调用大模型更新理解与 Prompt。"""
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "请先上传并完成首次澄清"}), 400
    user_answers = body.get("user_answers") or {}
    prompt_revision_instruction = (body.get("prompt_revision_instruction") or "").strip()
    has_answer = any(str(v).strip() for v in user_answers.values())
    if not has_answer and not prompt_revision_instruction:
        return jsonify({"error": "请先填写至少一条补充说明，或填写 Prompt 修改要求"}), 400

    previous = job.get("clarified") or {}
    questions = (
        previous.get("clarifying_questions")
        or job.get("clarifying_questions_snapshot")
        or []
    )
    if not questions and not prompt_revision_instruction:
        return jsonify({"error": "当前没有待澄清问题，无需提交补充"}), 400

    try:
        mi, vendor_info = _resolve_clarify_model(body, job)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    vendor = mi.get("vendor", "google")
    model_key = mi.get("key", body.get("model_key") or job.get("clarify_model_key") or "gemini-3.5-flash")
    _save_clarify_model_prefs(job, body, model_key)
    api_key = _resolve_presented_api_key(vendor, body.get("api_key"))
    if not api_key:
        return jsonify({"error": f"请填写 {vendor_info.get('name', vendor)} 的 API Key 或演示口令"}), 400

    path = Path(job["path"])
    file_data = read_rows(path)
    columns = job.get("columns") or file_data["columns"]
    _, preview, total = preview_from_table(
        file_data["columns"], file_data["rows"], job.get("data_rows"), n=5
    )
    preview_for_llm = _compact_preview_rows(preview, max_rows=3, max_cell_chars=180)
    fm_draft, _ = validate_field_mapping(
        columns,
        body.get("field_mapping") or previous.get("field_mapping") or {},
    )
    tmpl = load_clarify_followup_template(TEMPLATES_DIR)
    prompt = render_clarify_followup_prompt(
        tmpl,
        task_description=job.get("task_description") or body.get("task_description") or "",
        user_output_expectation=job.get("user_output_expectation") or body.get("user_output_expectation") or "",
        file_path=str(path),
        columns_list=columns,
        preview_rows=preview_for_llm,
        previous_clarified=previous,
        user_answers=user_answers,
        clarifying_questions=questions,
        current_field_mapping=fm_draft,
        current_prompt=body.get("prompt_template") or previous.get("suggested_prompt") or "",
        prompt_revision_instruction=prompt_revision_instruction,
        rubric_text=job.get("rubric_text") or "",
    )
    try:
        raw = call_llm(mi["provider"], mi["model_id"], api_key, prompt, 0.15, mi.get("base_url"))
    except Exception as e:
        return jsonify({"error": f"补充澄清失败: {e}"}), 500

    parsed = parse_model_json(raw)
    parsed, repaired = _repair_suggested_prompt_with_llm(
        parsed,
        provider=mi["provider"],
        model_id=mi["model_id"],
        api_key=api_key,
        base_url=mi.get("base_url"),
        task_description=job.get("task_description") or body.get("task_description") or "",
        user_output_expectation=job.get("user_output_expectation") or body.get("user_output_expectation") or "",
        columns=columns,
        field_mapping_hint=fm_draft,
    )
    clarified, mapping_warnings = _clarified_from_parsed(parsed, columns)
    if repaired:
        clarified["mapping_warnings"] = list(clarified.get("mapping_warnings") or []) + ["首轮未产出 Prompt，已由模型自动补写。"]
    col_stats = job.get("column_stats") or compute_column_stats(columns, file_data["rows"])
    clarified["annotation_column_candidates"] = suggest_annotation_columns(
        columns, file_data["rows"], columns_stats=col_stats
    )

    # 第二轮：面向用户的最终解读（补充完成后必须再解读一轮）
    interpret_summary = ""
    interpret_checklist: list = []
    try:
        itmpl = load_clarify_interpret_template(TEMPLATES_DIR)
        iprompt = render_clarify_interpret_prompt(
            itmpl,
            clarified=clarified,
            user_answers=user_answers,
            clarifying_questions=questions,
        )
        raw_interpret = call_llm(mi["provider"], mi["model_id"], api_key, iprompt, 0.1, mi.get("base_url"))
        interp_parsed = parse_model_json(raw_interpret)
        if not interp_parsed.get("parse_error"):
            interpret_summary = interp_parsed.get("interpretation_summary") or ""
            interpret_checklist = interp_parsed.get("checklist") or []
            clarified["remaining_risks"] = interp_parsed.get("remaining_risks") or []
        else:
            interpret_summary = (raw_interpret or "")[:2000]
    except Exception as e:
        interpret_summary = f"（解读轮次调用失败: {e}，请根据上方配置人工核对）"

    clarified["interpretation_summary"] = interpret_summary
    clarified["interpretation_checklist"] = interpret_checklist
    job["clarified"] = clarified
    job["clarifying_questions_snapshot"] = clarified.get("clarifying_questions") or []
    job["user_answers"] = user_answers
    job["prompt_revision_instruction"] = prompt_revision_instruction
    job["status"] = "clarifying"
    return jsonify({
        "status": "clarifying",
        "job_id": job_id,
        "clarified": clarified,
        "columns": columns,
        "column_stats": col_stats,
        "data_rows": total or job.get("data_rows"),
        "reclarified": True,
        "interpretation_summary": interpret_summary,
        "model_used": {
            "vendor": vendor,
            "vendor_name": vendor_info.get("name", vendor),
            "model_key": model_key,
            "model_name": mi["name"],
        },
    })


@app.route("/api/confirm", methods=["POST"])
def confirm():
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "无效 job_id"}), 400
    columns = job.get("columns") or []
    fm_in = body.get("field_mapping") or (job.get("clarified") or {}).get("field_mapping", {})
    field_mapping, more_warnings = validate_field_mapping(columns, fm_in)
    job["field_mapping"] = field_mapping
    job["mapping_warnings"] = more_warnings
    job["prompt_template"] = body.get("prompt_template") or (job.get("clarified") or {}).get("suggested_prompt", "")
    job["user_answers"] = body.get("user_answers") or job.get("user_answers") or {}
    job["prompt_revision_instruction"] = (
        body.get("prompt_revision_instruction")
        or job.get("prompt_revision_instruction")
        or ""
    )
    task_type_raw = (body.get("task_type") or "").strip()
    fallback_task_type = (
        task_type_raw
        or (job.get("clarified") or {}).get("task_type")
        or (job.get("task_description") or "").split("，")[0].split("。")[0].strip()
        or Path(job.get("filename") or "未命名任务").stem
    )
    job["task_type"] = fallback_task_type
    job["output_format"] = body.get("output_format", "") or (job.get("clarified") or {}).get("output_format", "")
    job["output_description"] = body.get("output_description", "") or (job.get("clarified") or {}).get("output_description", "")
    job["user_output_expectation"] = body.get("user_output_expectation") or job.get("user_output_expectation", "")
    job["strategies"] = normalize_strategies(body.get("strategies"))
    job["status"] = "confirmed"
    if isinstance(job.get("clarified"), dict):
        job["clarified"]["task_type"] = job["task_type"]
    file_kind = job.get("file_kind") or ("xlsx" if str(job.get("path", "")).lower().endswith(".xlsx") else "csv")
    embedded_cols = job.get("embedded_image_columns") or []
    sample_rows: list[list[str]] = []
    try:
        src = Path(job["path"])
        if src.is_file():
            sample_rows = (read_rows(src).get("rows") or [])[:20]
    except Exception:
        pass
    vision_info = assess_job_vision(
        field_mapping,
        file_kind=file_kind,
        embedded_image_columns=embedded_cols,
        columns=columns,
        sample_rows=sample_rows,
    )
    job["vision_required"] = vision_info["vision_required"]
    job["primary_content_mode"] = vision_info["primary_content_mode"]
    field_mapping["vision_required"] = vision_info["vision_required"]
    field_mapping["primary_content_mode"] = vision_info["primary_content_mode"]
    job["field_mapping"] = field_mapping
    questions = (job.get("clarified") or {}).get("clarifying_questions") or []
    supplement_warning = ""
    if questions and not job.get("user_answers"):
        supplement_warning = (
            f"模型提出了 {len(questions)} 个澄清问题；若未填写，请先点「提交补充，AI 重新理解」。"
        )
        more_warnings = list(more_warnings) + [supplement_warning]
    gt_stats = {"ground_truth_column": field_mapping.get("ground_truth"), "filled_rows": 0, "total_rows": 0}
    try:
        src = Path(job["path"])
        if src.is_file():
            fd = read_rows(src)
            gt_stats = count_ground_truth_rows(fd["columns"], fd["rows"], field_mapping)
    except Exception:
        pass
    if vision_info["vision_required"]:
        more_warnings = list(more_warnings) + [
            "本任务主内容为图片作答（URL 或 Excel 内嵌图），跑批将启用多模态传图；"
            "请在第三步仅选择支持看图的模型（如 Gemini 3.5、GPT-5.5、豆包 Seed、Qwen3.6 等多模态模型）。"
        ]
    return jsonify({
        "status": "confirmed",
        "job_id": job_id,
        "prompt_template": job["prompt_template"],
        "prompt_revision_instruction": job.get("prompt_revision_instruction", ""),
        "user_supplement_warning": supplement_warning,
        "mapping_warnings": more_warnings,
        "ground_truth_column": field_mapping.get("ground_truth"),
        "ground_truth_stats": gt_stats,
        "strategies": job.get("strategies"),
        "vision_required": vision_info["vision_required"],
        "primary_content_mode": vision_info["primary_content_mode"],
        "embedded_image_columns": embedded_cols,
    })


@app.route("/api/prepare", methods=["POST"])
def prepare():
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "无效 job_id"}), 400
    if job.get("status") != "confirmed" and not job.get("field_mapping"):
        return jsonify({"error": "请先确认字段映射（/api/confirm）"}), 400
    fm = job.get("field_mapping") or {}
    src = Path(job["path"])
    file_data = read_rows(src)
    info = backup_and_prepare(
        src,
        BACKUP_DIR,
        PREPARED_DIR,
        fm,
        columns=file_data["columns"],
        rows=file_data["rows"],
    )
    job["backup_path"] = info["backup_path"]
    job["prepared_path"] = info["prepared_path"]
    job["status"] = "ready"
    gt_stats = count_ground_truth_rows(file_data["columns"], file_data["rows"], fm)
    return jsonify({
        "status": "ready",
        "job_id": job_id,
        "ground_truth_column": fm.get("ground_truth"),
        "ground_truth_stats": gt_stats,
        **info,
    })


def _probe_prompt_rows(
    *,
    columns: list[str],
    rows: list[list[str]],
    cfg: dict,
    model_info: dict,
    sample_n: int = 3,
) -> dict:
    """用单模型试跑若干行，检查输出是否可解析、可与标注比对（与正式跑批同一套解析）。"""
    api_keys = cfg.get("api_keys") or {}
    default_api_key = cfg.get("api_key", "")
    vendor = model_info.get("vendor", "google")
    api_key = api_keys.get(vendor) or default_api_key
    if not api_key:
        raise ValueError(f"缺少 {VENDORS.get(vendor, {}).get('name', vendor)} 的 API Key")

    field_mapping = finalize_field_mapping(cfg.get("field_mapping") or {})
    profile = infer_scoring_profile(columns, field_mapping)
    strategies = normalize_strategies(cfg.get("strategies") or {})
    samples: list[dict] = []
    parseable = 0
    comparable = 0

    for i, row_values in enumerate(rows[:sample_n]):
        item = _process_one_row(
            i,
            row_values,
            columns=columns,
            cfg=cfg,
            model_info=model_info,
            api_key=api_key,
        )
        item["response_preview"] = (item.get("response") or "")[:280]
        if item.get("status") == "ok":
            by_col = item.get("ground_truth_by_column") or {}
            has_pred = any(
                (by_col.get(c) or {}).get("predicted") is not None
                for c in list_ground_truth_columns(field_mapping)
            )
            if has_pred or item.get("predicted"):
                parseable += 1
            if item.get("correct") is not None:
                comparable += 1
        samples.append({
            "row": item.get("row", i + 1),
            "status": item.get("status", "error"),
            "predicted": item.get("predicted"),
            "ground_truth": item.get("ground_truth"),
            "correct": item.get("correct"),
            "response_preview": item.get("response_preview"),
            "error": item.get("response") if item.get("status") == "error" else None,
            "is_multi_point": profile.get("is_multi_point"),
        })

    n = len(rows[:sample_n])
    return {
        "rows_probed": n,
        "rows_parseable": parseable,
        "rows_comparable": comparable,
        "samples": samples,
        "is_multi_point": profile.get("is_multi_point"),
    }


@app.route("/api/prompt_engineering_preview", methods=["POST"])
def prompt_engineering_preview():
    """返回当前列映射下，Prompt Engineering 将追加到 Prompt 末尾的文案（不落库）。"""
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id")
    fm_raw = body.get("field_mapping") or {}
    job = _get_job(job_id) if job_id else None
    if job:
        cols = job.get("columns") or []
        if cols and fm_raw:
            fm, _ = validate_field_mapping(cols, fm_raw)
        else:
            fm = finalize_field_mapping(fm_raw)
    else:
        fm = finalize_field_mapping(fm_raw)
    strategies = normalize_strategies(body.get("strategies") or {"enabled": True, "json_output": True})
    append = preview_prompt_engineering_append(fm, strategies)
    return jsonify({"append": append or ""})


@app.route("/api/validate_prompt", methods=["POST"])
def validate_prompt():
    """第 2→3 步前：试渲染 Prompt；可选试跑 3 行检查输出能否与人工标注比对。"""
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "无效 job_id"}), 400
    if not job.get("prepared_path"):
        return jsonify({"error": "请先点击「确认并准备数据」"}), 400

    _apply_field_mapping_to_job(job, body)
    prompt = (body.get("prompt_template") or job.get("prompt_template") or "").strip()
    job["prompt_template"] = prompt
    fm = job.get("field_mapping") or {}

    file_data = read_rows(_resolve_data_path(job))
    columns = file_data["columns"]
    rows = file_data["rows"]
    rubric_extra = {}
    if (job.get("rubric_text") or "").strip():
        rubric_extra["rubric_content"] = job["rubric_text"].strip()
    strategies = normalize_strategies(body.get("strategies") or job.get("strategies") or {})
    _apply_strategies_to_job(job, body)
    render_errors = validate_prompt_renders(
        prompt,
        columns=columns,
        rows=rows,
        field_mapping=fm,
        sample_n=3,
        extra_vars=rubric_extra or None,
    )
    pe_append = preview_prompt_engineering_append(fm, strategies) if strategies.get("enabled") else ""
    sample_rendered = None
    if rows and prompt.strip():
        rd = row_to_dict(columns, rows[0])
        sample_rendered = render_row_prompt(prompt, rd, fm, extra_vars=rubric_extra or None)
        if pe_append:
            sample_rendered = sample_rendered.rstrip() + "\n\n" + pe_append
    gt_stats = count_ground_truth_rows(columns, rows, fm)

    out: dict = {
        "prompt_render_ok": len(render_errors) == 0,
        "render_errors": render_errors,
        "ground_truth_column": fm.get("ground_truth"),
        "ground_truth_stats": gt_stats,
        "strategies": strategies,
        "strategies_summary": describe_active_strategies(strategies),
        "prompt_engineering_enabled": bool(strategies.get("enabled") and pe_append),
        "prompt_engineering_append": pe_append,
        "sample_rendered_tail": (sample_rendered or "")[-1200:] if sample_rendered else None,
        "llm_probe": None,
        "ready_for_batch": len(render_errors) == 0,
    }

    if body.get("run_llm_probe") and out["prompt_render_ok"]:
        model_key = (body.get("model_key") or "").strip()
        if model_key and model_key in MODELS:
            specs = [dict(MODELS[model_key])]
        else:
            specs = resolve_model_specs(
                body.get("models") or ["gemini-3.5-flash"],
                body.get("custom_models"),
            )
        if not specs:
            return jsonify({"error": "试跑校验需要指定一个模型"}), 400
        api_keys = resolve_api_keys(body, specs[:1])
        key_err = validate_api_keys_for_specs(specs[:1], api_keys)
        if key_err:
            return jsonify({"error": key_err}), 400
        cfg = {
            "task_description": job.get("task_description", ""),
            "api_key": (body.get("api_key") or "").strip(),
            "api_keys": api_keys,
            "prompt_template": prompt,
            "field_mapping": fm,
            "temperature": 0,
            "strategies": strategies,
            "file_kind": job.get("file_kind") or "csv",
            "image_source_path": job.get("path") or "",
            "embedded_image_columns": job.get("embedded_image_columns") or [],
            "vision_required": job.get("vision_required"),
            "rubric_text": (job.get("rubric_text") or "").strip(),
        }
        try:
            probe = _probe_prompt_rows(
                columns=columns,
                rows=rows,
                cfg=cfg,
                model_info=specs[0],
                sample_n=int(body.get("probe_rows") or 1),
            )
            out["llm_probe"] = probe
            gt_col = fm.get("ground_truth")
            if probe.get("is_multi_point") and probe["rows_comparable"] == 0 and probe["rows_probed"] > 0:
                out["ready_for_batch"] = False
                pts = list_ground_truth_columns(fm)
                point_cols = [c for c in pts if "采分点" in str(c)]
                out["probe_warning"] = (
                    f"已试跑 {probe['rows_probed']} 行，但未能将模型 JSON 对齐到各采分点列。"
                    f"请要求 JSON 键名与表头完全一致（如 {', '.join(point_cols[:5])}、总分），"
                    "勿仅用 ①② 或采分点① 等别名。"
                )
            elif gt_col and probe["rows_comparable"] == 0 and probe["rows_probed"] > 0:
                out["ready_for_batch"] = False
                out["probe_warning"] = (
                    f"已试跑 {probe['rows_probed']} 行，但未能从模型输出解析出可与标注列「{gt_col}」比对的字段。"
                    "请调整 Prompt，要求 JSON 输出 is_correct / prediction（0/1 与标注一致）。"
                )
            elif gt_col and probe["rows_parseable"] == 0:
                out["ready_for_batch"] = False
                out["probe_warning"] = "模型有返回，但输出格式无法解析，请收紧 Prompt 为固定 JSON。"
        except Exception as e:
            out["llm_probe"] = {"error": str(e)}
            out["ready_for_batch"] = False

    job["prompt_validated"] = out.get("ready_for_batch", False)
    return jsonify(out)


@app.route("/api/run_sample", methods=["POST"])
def run_sample():
    """单元测试：默认 5 行，可多模型对比。"""
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "请先上传并准备数据"}), 400
    if not job.get("prepared_path"):
        return jsonify({"error": "请先调用 /api/prepare 备份并生成规范数据"}), 400
    _apply_field_mapping_to_job(job, body)

    model_keys = body.get("models") or []
    custom_models = body.get("custom_models") or []
    model_specs = resolve_model_specs(model_keys, custom_models)
    if not model_specs:
        return jsonify({"error": "请至少选择一个预设模型，或添加自定义模型 ID"}), 400
    api_keys = resolve_api_keys(body, model_specs)
    key_err = validate_api_keys_for_specs(model_specs, api_keys)
    if key_err:
        return jsonify({"error": key_err}), 400
    sample_n = int(body.get("rows") or 5)
    if body.get("prompt_template"):
        job["prompt_template"] = body.get("prompt_template")

    run_id = uuid.uuid4().hex[:10]
    cfg = _run_config_from_body(job, body, model_specs, mode="sample")
    cfg["max_rows"] = sample_n
    cfg["api_keys"] = api_keys
    RUNS[run_id] = _init_run(run_id, job_id, job, cfg)

    def _worker():
        run = RUNS[run_id]
        try:
            file_data = read_rows(_resolve_data_path(job))
            rows = select_sample_rows(file_data["rows"], max_rows=sample_n)
            _run_rows_for_models(run, columns=file_data["columns"], rows=rows, cfg=cfg, model_specs=model_specs)
        except Exception:
            _mark_run_finished(run, "error", traceback.format_exc())

    est_sec = _estimate_run_seconds(job, cfg, model_specs)
    if est_sec is not None:
        RUNS[run_id]["estimated_total_sec_hint"] = est_sec

    threading.Thread(target=_worker, daemon=True).start()
    return jsonify({
        "run_id": run_id,
        "status": "queued",
        "mode": "sample",
        "rows": sample_n,
        "concurrency": cfg.get("concurrency"),
        "estimated_total_sec": est_sec,
        "estimated_total_label": _format_duration_sec(est_sec),
    })


@app.route("/api/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "缺少 file 字段"}), 400
    f = request.files["file"]
    if not f or not f.filename:
        return jsonify({"error": "未选择文件"}), 400
    ext = Path(f.filename).suffix.lower()
    if ext not in (".xlsx", ".csv"):
        return jsonify({"error": "仅支持 .xlsx / .csv"}), 400

    job_id = uuid.uuid4().hex[:10]
    safe = f"{job_id}_{_safe_name(f.filename)}"
    path = UPLOAD_DIR / safe
    f.save(path)

    try:
        columns, data_rows = inspect_file(path)
        file_data = read_rows(path)
        col_stats = compute_column_stats(file_data["columns"], file_data["rows"])
    except Exception as e:
        path.unlink(missing_ok=True)
        return jsonify({"error": f"解析失败: {e!s}"}), 400

    file_kind = "xlsx" if ext == ".xlsx" else "csv"
    embedded_cols: list[str] = []
    if file_kind == "xlsx":
        try:
            embedded_cols = detect_embedded_image_columns(path, columns)
        except Exception:
            embedded_cols = []

    JOBS[job_id] = {
        "job_id": job_id,
        "filename": f.filename,
        "saved_name": safe,
        "columns": columns,
        "data_rows": data_rows,
        "column_stats": col_stats,
        "path": str(path),
        "file_kind": file_kind,
        "embedded_image_columns": embedded_cols,
    }

    vision_hint = None
    if embedded_cols:
        vision_hint = (
            f"检测到 Excel 内嵌图片列：{'、'.join(embedded_cols)}。"
            "跑批时将把图片送入多模态模型，请选用支持看图的模型（如 Gemini / GPT-4o）。"
        )
    elif file_kind == "csv":
        low_cols = [str(c).lower() for c in columns]
        if any("image" in c or "图" in c for c in low_cols):
            vision_hint = (
                "当前为 CSV：无法携带 Excel 单元格内嵌图。"
                "若主输入是切图，请改上传 .xlsx（如 sample_governed.xlsx）。"
            )

    return jsonify({
        "job_id": job_id,
        "filename": f.filename,
        "columns": columns,
        "data_rows": data_rows,
        "column_count": len(columns),
        "column_stats": col_stats,
        "file_kind": file_kind,
        "embedded_image_columns": embedded_cols,
        "vision_hint": vision_hint,
    })


@app.route("/api/upload_rubric", methods=["POST"])
def upload_rubric():
    """上传评分标准文本（.txt / .md），表内无规则列时使用。"""
    job_id = request.form.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "请先上传数据文件"}), 400
    if "file" not in request.files or not request.files["file"]:
        return jsonify({"error": "请选择 .txt 或 .md 文件"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "未选择文件"}), 400
    ext = Path(f.filename).suffix.lower()
    if ext not in (".txt", ".md"):
        return jsonify({"error": "评分标准仅支持 .txt / .md"}), 400
    raw = f.read()
    if len(raw) > 512 * 1024:
        return jsonify({"error": "评分标准文件过大（最大 512KB）"}), 400
    text = ""
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if not text.strip():
        return jsonify({"error": "无法解码或文件为空，请使用 UTF-8 / GBK 编码的文本"}), 400
    text = text.strip()[:50000]
    job["rubric_text"] = text
    job["rubric_filename"] = f.filename
    return jsonify({
        "ok": True,
        "job_id": job_id,
        "rubric_filename": f.filename,
        "rubric_chars": len(text),
        "rubric_preview": text[:400] + ("…" if len(text) > 400 else ""),
    })


@app.route("/api/delete_rubric", methods=["POST"])
def delete_rubric():
    """删除当前任务已上传的评分标准。"""
    body = request.get_json(silent=True) or {}
    job_id = body.get("job_id") or request.form.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "请先上传数据文件"}), 400
    job.pop("rubric_text", None)
    job.pop("rubric_filename", None)
    return jsonify({"ok": True, "job_id": job_id})


@app.route("/api/run", methods=["POST"])
def start_run():
    body = request.get_json(silent=True) or {}

    job_id = body.get("job_id")
    job = _get_job(job_id)
    if not job:
        return jsonify({"error": "请先上传文件"}), 400

    task_desc = (body.get("task_description") or job.get("task_description") or "").strip()
    if not task_desc:
        return jsonify({"error": "请填写任务说明（描述你想用 AI 做什么）"}), 400
    if not job.get("prepared_path"):
        return jsonify({"error": "全量评测前请先完成「确认规则」并执行数据准备（/api/prepare）"}), 400
    _apply_field_mapping_to_job(job, body)

    model_keys = body.get("models") or []
    custom_models = body.get("custom_models") or []
    model_specs = resolve_model_specs(model_keys, custom_models)
    if not model_specs:
        return jsonify({"error": "请至少选择一个预设模型，或添加自定义模型 ID"}), 400
    api_keys = resolve_api_keys(body, model_specs)
    key_err = validate_api_keys_for_specs(model_specs, api_keys)
    if key_err:
        return jsonify({"error": key_err}), 400

    if body.get("prompt_template"):
        job["prompt_template"] = body.get("prompt_template")

    run_id = uuid.uuid4().hex[:10]
    cfg = _run_config_from_body(job, body, model_specs, mode="batch")
    cfg["api_keys"] = api_keys
    cfg["task_description"] = task_desc

    RUNS[run_id] = _init_run(run_id, job_id, job, cfg)

    est_sec = _estimate_run_seconds(job, cfg, model_specs)
    if est_sec is not None:
        RUNS[run_id]["estimated_total_sec_hint"] = est_sec

    t = threading.Thread(target=_batch_worker, args=(run_id,), daemon=True)
    t.start()

    return jsonify({
        "run_id": run_id,
        "status": "queued",
        "estimated_total_sec": est_sec,
        "estimated_total_label": _format_duration_sec(est_sec),
    })


@app.route("/api/status/<run_id>")
def get_status(run_id):
    run = RUNS.get(run_id)
    if not run:
        return jsonify({"error": "未找到该运行记录"}), 404
    payload = {
        "run_id": run_id,
        "status": run["status"],
        "progress": round(run["progress"], 4),
        "processed_in_model": run["processed_in_model"],
        "total_rows": run["total_rows"],
        "total_models": run["total_models"],
        "current_model": run["current_model"],
        "current_model_idx": run["current_model_idx"],
        "model_progress": run.get("model_progress") or [],
        "concurrency": run.get("concurrency"),
        "cancel_requested": bool(run.get("cancel_requested")),
        "error": run.get("error"),
    }
    payload.update(_run_timing_fields(run))
    return jsonify(payload)


@app.route("/api/cancel/<run_id>", methods=["POST"])
def cancel_run(run_id):
    run = RUNS.get(run_id)
    if not run:
        return jsonify({"error": "未找到该运行记录"}), 404
    if run["status"] not in ("queued", "running"):
        return jsonify({
            "error": "当前状态无法中止",
            "status": run["status"],
        }), 400
    run["cancel_requested"] = True
    return jsonify({"ok": True, "run_id": run_id, "message": "已请求中止，正在等待进行中的请求结束…"})


@app.route("/api/results/<run_id>")
def get_results(run_id):
    run = RUNS.get(run_id)
    if not run:
        return jsonify({"error": "未找到该运行记录"}), 404
    if run["status"] not in ("done", "error", "cancelled"):
        return jsonify({"error": "任务尚未完成", "status": run["status"]}), 400

    if run["status"] == "error":
        return jsonify({"status": "error", "error": run.get("error", "未知错误")}), 200

    job = JOBS.get(run["job_id"], {})
    fm = finalize_field_mapping(
        (run.get("config") or {}).get("field_mapping") or job.get("field_mapping") or {}
    )
    gt_col = fm.get("primary_ground_truth") or fm.get("ground_truth")
    gt_cols = list_ground_truth_columns(fm)
    models_out = []
    all_samples = []
    best_accuracy = None
    best_model = None
    raw_model_results = list(run.get("model_results", []))
    vote_result = _build_ensemble_vote_result(raw_model_results, fm, gt_cols)
    result_entries = ([vote_result] if vote_result else []) + raw_model_results

    for mr in result_entries:
        ok_count = sum(1 for r in mr["responses"] if r["status"] == "ok")
        err_count = mr["errors"]
        metrics = mr.get("metrics") or metrics_for_responses(
            mr.get("responses", []),
            ground_truth_column=gt_col,
            ground_truth_columns=gt_cols,
            field_mapping=fm,
        )
        acc = metrics.get("accuracy")
        entry = {
            "name": mr["model_name"],
            "model_id": mr.get("model_id", ""),
            "rows_ok": ok_count,
            "rows_error": err_count,
            "time_sec": mr.get("duration_sec", 0),
            "metrics": metrics,
            "accuracy": acc,
            "accuracy_pct": metrics.get("accuracy_pct"),
            "synthetic": bool(mr.get("synthetic")),
        }
        responses = mr.get("responses", [])
        total_row_time = sum(float(r.get("time_sec") or 0) for r in responses)
        total_cost = sum(float(r.get("estimated_cost_usd") or 0) for r in responses)
        total_cost_cny = sum(float(r.get("estimated_cost_cny") or 0) for r in responses)
        cost_estimated_rows = sum(1 for r in responses if r.get("cost_estimated"))
        api_calls = sum(int(r.get("api_calls") or 1) for r in responses)
        prompt_tokens = sum(int(r.get("prompt_tokens") or 0) for r in responses)
        completion_tokens = sum(int(r.get("completion_tokens") or 0) for r in responses)
        reasoning_tokens = sum(int(r.get("reasoning_tokens") or 0) for r in responses)
        stability_rows = [r for r in responses if r.get("stability_consistent") is not None]
        stability_ok = sum(1 for r in stability_rows if r.get("stability_consistent") is True)
        entry.update({
            "avg_time_sec_per_row": round(total_row_time / ok_count, 2) if ok_count else None,
            "avg_cost_usd_per_row": round(total_cost / ok_count, 6) if ok_count else None,
            "total_estimated_cost_usd": round(total_cost, 6),
            "avg_cost_cny_per_row": round(total_cost_cny / ok_count, 6) if ok_count else None,
            "total_estimated_cost_cny": round(total_cost_cny, 6),
            "cost_estimated_rows": cost_estimated_rows,
            "prompt_tokens": prompt_tokens,
            "completion_tokens": completion_tokens,
            "reasoning_tokens": reasoning_tokens,
            "api_calls": api_calls,
            "api_ok_calls": sum(int(r.get("api_calls") or 1) for r in responses if r.get("status") == "ok"),
            "stability_checked_rows": len(stability_rows),
            "stability_consistent_rows": stability_ok,
            "stability_consistency_pct": round(stability_ok / len(stability_rows) * 100, 2) if stability_rows else None,
        })
        models_out.append(entry)
        if acc is not None and (best_accuracy is None or acc > best_accuracy):
            best_accuracy = acc
            best_model = mr["model_name"]
        for r in mr["responses"][:50]:
            if not isinstance(r, dict):
                continue
            pred = r.get("predicted")
            pred_s = "" if pred is None else str(pred)[:80]
            gt = r.get("ground_truth")
            gt_s = "" if gt is None else str(gt)[:80]
            match = r.get("correct")
            match_label = "—"
            if match is True:
                match_label = "✓"
            elif match is False:
                match_label = "✗"
            all_samples.append({
                "row": r["row"],
                "model": mr["model_name"],
                "status": r["status"],
                "response": (r.get("response") or "")[:500],
                "predicted": pred_s,
                "ground_truth": gt_s,
                "correct": match,
                "match_label": match_label,
                "match_explanation": (r.get("match_explanation") or "")[:300],
                "estimated_cost_usd": r.get("estimated_cost_usd"),
                "estimated_cost_cny": r.get("estimated_cost_cny"),
                "stability_consistent": r.get("stability_consistent"),
                "time_sec": r.get("time_sec", 0),
            })

    total = run["total_rows"]
    real_models_out = [m for m in models_out if not m.get("synthetic")]
    duration = sum(m["time_sec"] for m in real_models_out)
    total_api_calls = sum(int(m.get("api_calls") or ((m.get("rows_ok") or 0) + (m.get("rows_error") or 0))) for m in real_models_out)
    total_api_ok = sum(int(m.get("api_ok_calls") or (m.get("rows_ok") or 0)) for m in real_models_out)
    total_estimated_cost = sum(float(m.get("total_estimated_cost_usd") or 0) for m in real_models_out)
    total_estimated_cost_cny = sum(float(m.get("total_estimated_cost_cny") or 0) for m in real_models_out)
    cost_estimated_rows = sum(int(m.get("cost_estimated_rows") or 0) for m in real_models_out)
    prompt_tokens = sum(int(m.get("prompt_tokens") or 0) for m in real_models_out)
    completion_tokens = sum(int(m.get("completion_tokens") or 0) for m in real_models_out)
    reasoning_tokens = sum(int(m.get("reasoning_tokens") or 0) for m in real_models_out)
    total_ok_rows = sum(int(m.get("rows_ok") or 0) for m in real_models_out)
    avg_time_values = [m.get("avg_time_sec_per_row") for m in real_models_out if m.get("avg_time_sec_per_row") is not None]
    stability_checked = sum(int(m.get("stability_checked_rows") or 0) for m in real_models_out)
    stability_consistent = sum(int(m.get("stability_consistent_rows") or 0) for m in real_models_out)
    best_entry = next((m for m in models_out if m.get("name") == best_model), None)
    best_met = (best_entry or {}).get("metrics") or {}
    is_multi = bool(best_met.get("is_multi_point"))
    has_ground_truth = bool(gt_cols) and any(
        (m.get("metrics") or {}).get("rows_with_ground_truth", 0) > 0
        for m in models_out
    )
    failure_buckets = _build_failure_buckets(raw_model_results, gt_cols)
    consensus_report = _build_consensus_report(raw_model_results)
    judgment_metric = best_met.get("judgment_metric_note") or (
        "多采分点：每行全部采分点与总分均与人工一致的比例"
        if is_multi
        else "模型预测与人工标注逐行比对的一致率"
    )
    result_model_specs = [
        {
            "name": mr.get("model_name"),
            "model_id": mr.get("model_id"),
            "provider": mr.get("provider") or "openai_compat",
        }
        for mr in raw_model_results
    ]
    summary = {
        "judgment_metric": judgment_metric,
        "judgment_metric_note": best_met.get("judgment_metric_note"),
        "evaluation_mode": best_met.get("evaluation_mode"),
        "evaluation_mode_label": best_met.get("evaluation_mode_label"),
        "is_multi_point": is_multi,
        "scoring_point_columns": best_met.get("scoring_point_columns"),
        "best_judgment_accuracy_pct": round(best_accuracy * 100, 2) if best_accuracy is not None else None,
        "best_model": best_model,
        "best_rows_correct": best_met.get("rows_correct"),
        "best_rows_evaluated": best_met.get("rows_evaluated"),
        "total_api_calls": total_api_calls,
        "total_api_ok": total_api_ok,
        "api_ok_pct": round(total_api_ok / total_api_calls * 100, 2) if total_api_calls else None,
        "avg_time_sec_per_row": round(sum(avg_time_values) / len(avg_time_values), 2) if avg_time_values else None,
        "avg_cost_usd_per_row": round(total_estimated_cost / total_ok_rows, 6) if total_ok_rows else None,
        "total_estimated_cost_usd": round(total_estimated_cost, 6),
        "avg_cost_cny_per_row": round(total_estimated_cost_cny / total_ok_rows, 6) if total_ok_rows else None,
        "total_estimated_cost_cny": round(total_estimated_cost_cny, 6),
        "cost_estimated_rows": cost_estimated_rows,
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "reasoning_tokens": reasoning_tokens,
        "stability_consistency_pct": round(stability_consistent / stability_checked * 100, 2) if stability_checked else None,
    }
    best_metric_pct = round(best_accuracy * 100, 2) if best_accuracy is not None else None
    automate_metric_pct = best_met.get("row_strict_accuracy_pct")
    if automate_metric_pct is None:
        automate_metric_pct = best_metric_pct
    if best_met.get("sequence_units_label"):
        summary["sequence_units_label"] = best_met.get("sequence_units_label")
    if best_met.get("row_strict_accuracy_pct") is not None:
        summary["row_strict_accuracy_pct"] = best_met.get("row_strict_accuracy_pct")
    summary["comparison_unit"] = best_met.get("comparison_unit") or fm.get("comparison_unit")
    summary["row_headline"] = fm.get("row_headline")
    readiness_report = _build_readiness_report(
        has_ground_truth=has_ground_truth,
        best_model=best_model,
        best_metric_pct=best_metric_pct,
        automate_metric_pct=automate_metric_pct,
        summary=summary,
        consensus=consensus_report,
        failure_buckets=failure_buckets,
    )
    sample_gate = _build_sample_gate(run, raw_model_results, readiness_report, consensus_report)
    harness_spec = _build_harness_spec(job, run.get("config") or {}, fm)
    payload = {
        "status": run["status"],
        "cancelled": run["status"] == "cancelled",
        "error": run.get("error") if run["status"] in ("error", "cancelled") else None,
        "run_id": run_id,
        "job_id": run.get("job_id"),
        "filename": job.get("filename", ""),
        "columns": job.get("columns") or [],
        "data_rows": job.get("data_rows") or total,
        "column_stats": job.get("column_stats") or {},
        "prepared_path": job.get("prepared_path"),
        "task_description": (run.get("config") or {}).get("task_description") or job.get("task_description", ""),
        "task_type": job.get("task_type") or (job.get("clarified") or {}).get("task_type") or "",
        "prompt_template": (run.get("config") or {}).get("prompt_template") or job.get("prompt_template", ""),
        "prompt_revision_instruction": job.get("prompt_revision_instruction", ""),
        "user_answers": job.get("user_answers") or {},
        "field_mapping": fm,
        "strategies": (run.get("config") or {}).get("strategies") or job.get("strategies") or {},
        "clarified": job.get("clarified") or {},
        "output_format": job.get("output_format") or "",
        "output_description": job.get("output_description") or "",
        "user_output_expectation": job.get("user_output_expectation") or "",
        "total_rows": total,
        "model_count": len(models_out),
        "duration_sec": round(duration, 1),
        "avg_time_sec_per_row": round(sum(avg_time_values) / len(avg_time_values), 2) if avg_time_values else None,
        "avg_cost_usd_per_row": round(total_estimated_cost / total_ok_rows, 6) if total_ok_rows else None,
        "total_estimated_cost_usd": round(total_estimated_cost, 6),
        "avg_cost_cny_per_row": round(total_estimated_cost_cny / total_ok_rows, 6) if total_ok_rows else None,
        "total_estimated_cost_cny": round(total_estimated_cost_cny, 6),
        "cost_estimated_rows": cost_estimated_rows,
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "reasoning_tokens": reasoning_tokens,
        "cost_estimation": build_cost_estimation_info(result_model_specs),
        "stability_checked_rows": stability_checked,
        "stability_consistent_rows": stability_consistent,
        "stability_consistency_pct": round(stability_consistent / stability_checked * 100, 2) if stability_checked else None,
        "ground_truth_column": gt_col,
        "best_model": best_model,
        "best_accuracy_pct": best_metric_pct,
        "best_judgment_accuracy_pct": best_metric_pct,
        "summary": summary,
        "harness_spec": harness_spec,
        "readiness_report": readiness_report,
        "sample_gate": sample_gate,
        "consensus_report": consensus_report,
        "failure_buckets": failure_buckets,
        "models": models_out,
        "samples": all_samples,
    }
    try:
        payload["snapshot_filename"] = _save_run_snapshot(run, payload)
        payload["snapshot_saved"] = True
    except Exception as e:
        payload["snapshot_saved"] = False
        payload["snapshot_error"] = str(e)
    return jsonify(payload)


@app.route("/api/snapshot/<run_id>")
def get_run_snapshot(run_id):
    run = RUNS.get(run_id)
    path = None
    if run and run.get("snapshot_path"):
        path = Path(run["snapshot_path"])
    else:
        candidate = SNAPSHOT_DIR / f"{run_id}.json"
        if candidate.is_file():
            path = candidate
    if not path or not path.is_file():
        return jsonify({"error": "未找到该运行快照，请先完成一次评测并打开结果页"}), 404
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"error": f"快照读取失败: {e}"}), 500
    return jsonify({"snapshot": data, "snapshot_filename": path.name})


@app.route("/api/export_results/<run_id>")
def export_results(run_id):
    run = RUNS.get(run_id)
    if not run:
        return jsonify({"error": "未找到该运行记录"}), 404
    if run.get("status") not in ("done", "cancelled"):
        return jsonify({"error": "运行尚未完成，暂不能导出", "status": run.get("status")}), 400
    job = JOBS.get(run.get("job_id"), {})
    try:
        file_data = read_rows(_resolve_data_path(job))
    except Exception as e:
        return jsonify({"error": f"读取准备数据失败: {e}"}), 500

    columns = list(file_data.get("columns") or [])
    rows = run.get("evaluated_rows") or file_data.get("rows") or []
    model_results = [mr for mr in (run.get("model_results") or []) if not mr.get("synthetic")]
    export_cols = list(columns)
    for mr in model_results:
        name = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", str(mr.get("model_name") or "model")).strip("_") or "model"
        for suffix in ("预测", "是否一致", "失败原因", "响应"):
            export_cols.append(f"{name}_{suffix}")

    out_rows: list[list[str]] = []
    for idx, row in enumerate(rows):
        out = list(row)
        for mr in model_results:
            resp_rows = mr.get("responses") or []
            r = _response_at_index(resp_rows, idx)
            correct = r.get("correct")
            if correct is True:
                correct_s = "一致"
            elif correct is False:
                correct_s = "不一致"
            else:
                correct_s = ""
            out.extend([
                "" if r.get("predicted") is None else str(r.get("predicted")),
                correct_s,
                str(r.get("response") or "") if r.get("status") != "ok" else "",
                str(r.get("response") or ""),
            ])
        out_rows.append(out)

    def generate():
        import io
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        writer.writerow(export_cols)
        yield "\ufeff".encode("utf-8") + buf.getvalue().encode("utf-8")
        for out in out_rows:
            buf.seek(0)
            buf.truncate(0)
            writer.writerow(out)
            yield buf.getvalue().encode("utf-8")

    stem = re.sub(r"[^A-Za-z0-9._-]", "_", Path(job.get("filename") or run_id).stem) or "export"
    ascii_filename = f"{stem}_export_{run_id}.csv"
    suffix = "_部分" if run.get("status") == "cancelled" else ""
    display_filename = f"{Path(job.get('filename') or run_id).stem}_AI跑批结果{suffix}_{run_id}.csv"
    return Response(
        generate(),
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{ascii_filename}\"; "
                f"filename*=UTF-8''{quote(display_filename, safe='')}"
            )
        },
    )


@app.route("/api/analyze_results", methods=["POST"])
def analyze_results():
    """评测结束后：大模型解读结果并给出优化策略建议。"""
    body = request.get_json(silent=True) or {}
    run_id = body.get("run_id")
    if not run_id or run_id not in RUNS:
        return jsonify({"error": "无效 run_id"}), 400
    run = RUNS[run_id]
    if run["status"] not in ("done", "cancelled"):
        return jsonify({"error": "评测尚未完成"}), 400

    job = JOBS.get(run["job_id"], {})
    model_key = (body.get("model_key") or job.get("clarify_model_key") or "gemini-3.5-flash").strip()
    if model_key not in MODELS:
        model_key = "gemini-3.5-flash"
    mi = dict(MODELS[model_key])
    api_keys = resolve_api_keys(body, [mi])
    api_key = api_keys.get(mi.get("vendor", "google"))
    if not api_key:
        return jsonify({"error": "请提供用于解读的 API Key"}), 400

    summary_rows = []
    for mr in run.get("model_results", []):
        met = mr.get("metrics") or {}
        summary_rows.append({
            "model": mr.get("model_name"),
            "accuracy_pct": met.get("accuracy_pct"),
            "rows_evaluated": met.get("rows_evaluated"),
            "rows_correct": met.get("rows_correct"),
            "rows_api_ok": met.get("rows_api_ok"),
            "rows_api_error": met.get("rows_api_error"),
            "message": met.get("message"),
        })

    row_examples: list[dict[str, Any]] = []
    for mr in run.get("model_results", [])[:1]:
        for r in (mr.get("responses") or [])[:15]:
            if r.get("status") != "ok":
                continue
            row_examples.append({
                "row": r.get("row"),
                "predicted": str(r.get("predicted") or "")[:260],
                "ground_truth": str(r.get("ground_truth") or "")[:260],
                "row_correct": r.get("correct"),
                "per_column_match": (r.get("match_explanation") or "")[:400],
            })

    cfg = run.get("config") or {}
    prompt = f"""你是自动化评测分析助手。根据一次多模型评测结果，先给出决策结论，再补充简短细节。

## 任务说明
{cfg.get("task_description") or job.get("task_description") or ""}

## 人工标注与评测方式
{json.dumps((cfg.get("field_mapping") or {}), ensure_ascii=False, indent=2)[:3000]}

## 当前 Prompt（节选）
{(cfg.get("prompt_template") or job.get("prompt_template") or "")[:2000]}

## 当前策略
{json.dumps(cfg.get("strategies") or job.get("strategies") or {}, ensure_ascii=False)}

## 评测结果摘要
{json.dumps(summary_rows, ensure_ascii=False, indent=2)}

## 逐行样例（含各标注列规则比对摘要 per_column_match：列名后 ✓/✗/?）
{json.dumps(row_examples, ensure_ascii=False, indent=2)}

请只输出 JSON（不要 markdown 围栏）。字段顺序重要：先结论后细节；findings 最多 3 条，suggestions 最多 3 条，每条 detail 不超过 40 字。
输出必须给人可读：verdict/next_action 要具体，避免“继续优化”这类空话。
strategy_recommendations 的字段必须严格对应第 2 步策略控件；只有 JSON 输出、One-shot、Few-shot、逐步核对、置信度会修改 Prompt，多模型和重复评测不修改 Prompt。
如果建议修改 Prompt，请在 prompt_revision 中给出可直接替换第 2 步「Prompt 模板」的完整模板，保留必要占位符（如 {{{{ primary_input }}}}）。
{{
  "verdict": "一句话结论（如：效果达标 / 需改 Prompt / 建议换模型）",
  "verdict_reason": "1-2 句原因，聚焦判断正确率与可比对性",
  "next_action": "下一步可执行动作（一句话）",
  "summary": "补充说明，不超过 2 句",
  "findings": ["要点1", "要点2"],
  "suggestions": [
    {{"title": "建议标题", "detail": "具体做法", "priority": "high|medium|low"}}
  ],
  "strategy_recommendations": {{
    "enabled": true,
    "prompt_engineering": false,
    "json_output": false,
    "one_shot": false,
    "few_shot": false,
    "chain_of_thought": false,
    "multi_model": false,
    "repeat_runs": false,
    "repeat_count": 2,
    "confidence_filter": false,
    "confidence_threshold": 0.8
  }},
  "prompt_revision": "若需修改 Prompt，给出完整或增补后的模板；无需修改则空字符串"
}}
"""
    try:
        raw = call_llm(
            mi["provider"],
            mi["model_id"],
            api_key,
            prompt,
            0.2,
            base_url=mi.get("base_url"),
        )
        parsed = parse_model_json(raw)
        parsed = normalize_result_analysis(parsed, raw)
        run["analysis"] = parsed
        return jsonify({"ok": True, "analysis": parsed})
    except Exception as e:
        return jsonify({"error": f"解读失败: {e}"}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5050"))
    print(f"\n  自动化评测: http://127.0.0.1:{port}")
    print(f"  上传目录: {UPLOAD_DIR}")
    key = os.environ.get("GEMINI_API_KEY")
    if key:
        print(f"  GEMINI_API_KEY: {key[:8]}...（已从环境变量读取）")
    else:
        print("  GEMINI_API_KEY: 未设置（请在界面填写或 set GEMINI_API_KEY=xxx）")
    print()
    app.run(host="127.0.0.1", port=port, debug=True)
